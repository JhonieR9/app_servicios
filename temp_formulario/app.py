from flask import Flask, request, render_template, flash, redirect, url_for, session, jsonify, send_file
import mysql.connector
import os
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
from functools import wraps
import bcrypt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

app = Flask(__name__)
app.config['SECRET_KEY'] = 'tu_clave_secreta_aqui_12345'

# Carpeta donde se guardarán las imágenes y archivos
app.config['UPLOAD_FOLDER'] = "static/uploads"
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB máximo
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Extensiones permitidas
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'doc', 'docx'}

# Contraseña de administrador (hash bcrypt de "admin123")
ADMIN_PASSWORD_HASH = bcrypt.hashpw("admin123".encode('utf-8'), bcrypt.gensalt())

# Tiempo de expiración de sesión (30 minutos)
SESSION_TIMEOUT = 30

def login_required(f):
    """Decorador para proteger rutas que requieren autenticación"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_logged_in' not in session:
            flash('Debe iniciar sesión como administrador para acceder', 'warning')
            return redirect(url_for('admin_login'))
        
        # Verificar timeout de sesión
        last_activity = session.get('last_activity')
        if last_activity:
            last_activity_time = datetime.fromisoformat(last_activity)
            if datetime.now() - last_activity_time > timedelta(minutes=SESSION_TIMEOUT):
                session.clear()
                flash('Su sesión ha expirado. Por favor inicie sesión nuevamente.', 'warning')
                return redirect(url_for('admin_login'))
        
        # Actualizar última actividad
        session['last_activity'] = datetime.now().isoformat()
        return f(*args, **kwargs)
    return decorated_function

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_db_connection():
    """Crea y retorna una nueva conexión a la base de datos"""
    try:
        return mysql.connector.connect(
            host="localhost",
            user="root",
            password="Jhonier18.",
            database="profiles_cv_db"
        )
    except mysql.connector.Error as err:
        print(f"Error de conexión: {err}")
        return None

def init_database():
    """Inicializar parámetros generales y crear tablas adicionales si no existen"""
    connection = get_db_connection()
    if not connection:
        return False
    
    cursor = connection.cursor()
    
    try:
        # Crear tabla de correo si no existe
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS correo_persona (
                id_correo INT AUTO_INCREMENT PRIMARY KEY,
                id_persona INT NOT NULL UNIQUE,
                correo VARCHAR(120),
                FOREIGN KEY (id_persona) REFERENCES personas(id_persona) ON DELETE CASCADE
            )
        """)
        
        # Crear tabla de servicios con valor por hora si no existe
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS servicios_persona (
                id_servicio INT AUTO_INCREMENT PRIMARY KEY,
                id_persona INT NOT NULL,
                categoria VARCHAR(50),
                descripcion TEXT NOT NULL,
                anios_experiencia DECIMAL(3,1),
                valor_hora DECIMAL(10,2),
                FOREIGN KEY (id_persona) REFERENCES personas(id_persona) ON DELETE CASCADE
            )
        """)
        
        print("Tablas creadas correctamente")
        
        # Verificar si ya hay parámetros
        cursor.execute("SELECT COUNT(*) FROM parametros_generales")
        count = cursor.fetchone()[0]
        
        if count == 0:
            # Insertar parámetros generales
            parametros = [
                ('tipo_documento',),
                ('genero',),
                ('horario',),
                ('dias_disponibles',),
                ('servicio_tipo',),
                ('servicios',)
            ]
            cursor.executemany("INSERT INTO parametros_generales (nombre_parametro) VALUES (%s)", parametros)
            
            # Obtener IDs de parámetros
            cursor.execute("SELECT id_parametro, nombre_parametro FROM parametros_generales")
            params = {row[1]: row[0] for row in cursor.fetchall()}
            
            # Insertar detalles de parámetros
            detalles = [
                # Tipo documento
                (params['tipo_documento'], 'Cédula de Ciudadanía', None),
                (params['tipo_documento'], 'Cédula de Extranjería', None),
                (params['tipo_documento'], 'Pasaporte', None),
                
                # Género
                (params['genero'], 'Masculino', None),
                (params['genero'], 'Femenino', None),
                (params['genero'], 'Otro', None),
                
                # Horarios
                (params['horario'], '8am-12pm', None),
                (params['horario'], '2pm-6pm', None),
                (params['horario'], '6pm-10pm', None),
                (params['horario'], '24 horas', None),
                
                # Días disponibles
                (params['dias_disponibles'], 'Entre Semana (Lunes-Viernes)', None),
                (params['dias_disponibles'], 'Toda la Semana', None),
                
                # Tipo de servicios
                (params['servicio_tipo'], 'Un servicio principal', None),
                (params['servicio_tipo'], 'Varios servicios', None),
            ]
            
            cursor.executemany(
                "INSERT INTO detalle_parametro (id_parametro, descripcion, valor_hora) VALUES (%s, %s, %s)",
                detalles
            )
            
            connection.commit()
            print("Parámetros inicializados correctamente")
        
        return True
        
    except Exception as e:
        print(f"Error al inicializar parámetros: {e}")
        connection.rollback()
        return False
    finally:
        cursor.close()
        connection.close()

@app.route('/')
def index():
    """Cargar formulario con parámetros dinámicos"""
    connection = get_db_connection()
    if not connection:
        return "Error de conexión a la base de datos"
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        # Obtener parámetros
        parametros = {}
        
        # Tipos de documento
        cursor.execute("""
            SELECT dp.id_detalle, dp.descripcion 
            FROM detalle_parametro dp
            JOIN parametros_generales pg ON dp.id_parametro = pg.id_parametro
            WHERE pg.nombre_parametro = 'tipo_documento'
        """)
        parametros['tipos_documento'] = cursor.fetchall()
        
        # Géneros
        cursor.execute("""
            SELECT dp.id_detalle, dp.descripcion 
            FROM detalle_parametro dp
            JOIN parametros_generales pg ON dp.id_parametro = pg.id_parametro
            WHERE pg.nombre_parametro = 'genero'
        """)
        parametros['generos'] = cursor.fetchall()
        
        # Horarios
        cursor.execute("""
            SELECT dp.id_detalle, dp.descripcion 
            FROM detalle_parametro dp
            JOIN parametros_generales pg ON dp.id_parametro = pg.id_parametro
            WHERE pg.nombre_parametro = 'horario'
        """)
        parametros['horarios'] = cursor.fetchall()
        
        # Días disponibles
        cursor.execute("""
            SELECT dp.id_detalle, dp.descripcion 
            FROM detalle_parametro dp
            JOIN parametros_generales pg ON dp.id_parametro = pg.id_parametro
            WHERE pg.nombre_parametro = 'dias_disponibles'
        """)
        parametros['dias'] = cursor.fetchall()
        
        # Tipos de servicios
        cursor.execute("""
            SELECT dp.id_detalle, dp.descripcion 
            FROM detalle_parametro dp
            JOIN parametros_generales pg ON dp.id_parametro = pg.id_parametro
            WHERE pg.nombre_parametro = 'servicio_tipo'
        """)
        parametros['servicios_tipo'] = cursor.fetchall()
        
        return render_template("formulario_actualizado.html", parametros=parametros)
        
    except Exception as e:
        return f"Error al cargar parámetros: {e}"
    finally:
        if connection and connection.is_connected():
            connection.close()

@app.route('/guardar', methods=['POST'])
def guardar():
    connection = None
    cursor = None
    
    try:
        print("DEBUG - Iniciando proceso de guardado")
        
        # Validar que el documento no exista ya
        numero_documento = request.form.get('numero_documento', '').strip()
        if numero_documento:
            connection = get_db_connection()
            if connection:
                cursor = connection.cursor()
                cursor.execute("SELECT COUNT(*) FROM personas WHERE numero_documento = %s", (numero_documento,))
                count = cursor.fetchone()[0]
                if count > 0:
                    flash(f'⚠️ El documento {numero_documento} ya está registrado en el sistema', 'warning')
                    cursor.close()
                    connection.close()
                    return redirect(url_for('index'))
                cursor.close()
                connection.close()
        
        # Validar archivo de foto de identificación
        if 'foto_identificacion' not in request.files:
            flash('No se seleccionó foto de identificación', 'danger')
            return redirect(url_for('index'))
        
        foto_file = request.files['foto_identificacion']
        
        if foto_file.filename == '':
            flash('No se seleccionó foto de identificación', 'danger')
            return redirect(url_for('index'))
        
        if not allowed_file(foto_file.filename):
            flash('Tipo de archivo no permitido para la foto', 'danger')
            return redirect(url_for('index'))
        
        # Generar nombres únicos para archivos
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Procesar foto de identificación
        foto_filename = f"foto_{timestamp}_{secure_filename(foto_file.filename)}"
        foto_path = os.path.join(app.config['UPLOAD_FOLDER'], foto_filename)
        foto_file.save(foto_path)
        
        # Procesar foto de antecedentes (opcional)
        antecedentes_file = request.files.get('foto_antecedentes')
        antecedentes_filename = None
        
        if antecedentes_file and antecedentes_file.filename and allowed_file(antecedentes_file.filename):
            antecedentes_filename = f"antecedentes_{timestamp}_{secure_filename(antecedentes_file.filename)}"
            antecedentes_path = os.path.join(app.config['UPLOAD_FOLDER'], antecedentes_filename)
            antecedentes_file.save(antecedentes_path)
        
        # Procesar archivo de recomendaciones (opcional)
        recomend_file = request.files.get('recomendaciones_archivo')
        recomend_filename = None
        
        if recomend_file and recomend_file.filename and allowed_file(recomend_file.filename):
            recomend_filename = f"recom_{timestamp}_{secure_filename(recomend_file.filename)}"
            recomend_path = os.path.join(app.config['UPLOAD_FOLDER'], recomend_filename)
            recomend_file.save(recomend_path)
        
        # Conectar a la base de datos
        connection = get_db_connection()
        if not connection:
            flash('Error al conectar con la base de datos', 'danger')
            return redirect(url_for('index'))
        
        cursor = connection.cursor()
        connection.start_transaction()
        
        # ====== 1. personas ======
        cursor.execute("""
            INSERT INTO personas 
            (id_tipo_documento, numero_documento, id_genero, nombre_completo, ciudad, codigo_dane, registrado_por)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            request.form['tipo_documento'],
            request.form['numero_documento'],
            request.form['genero'],
            request.form['nombre_completo'],
            request.form['ciudad'],
            request.form['codigo_dane'],
            request.form['nombre_completo']  # Guardar el nombre de quien se registra
        ))
        
        id_persona = cursor.lastrowid
        print(f"DEBUG - Persona creada con ID: {id_persona}")
        
        # ====== 2. telefono_persona ======
        cursor.execute("""
            INSERT INTO telefono_persona (id_persona, telefono)
            VALUES (%s, %s)
        """, (id_persona, request.form['celular']))
        
        # ====== 2b. correo_persona (opcional) ======
        correo = request.form.get('correo', '').strip()
        if correo:
            cursor.execute("""
                INSERT INTO correo_persona (id_persona, correo)
                VALUES (%s, %s)
            """, (id_persona, correo))
        
        # ====== 3. servicios_persona (con valor por hora y categoría) ======
        servicios_categoria = request.form.getlist('servicio_categoria[]')
        servicios_desc = request.form.getlist('habilidad_desc[]')
        servicios_exp = request.form.getlist('habilidad_exp[]')
        servicios_valor = request.form.getlist('habilidad_valor[]')
        tiene_ayudante_list = request.form.getlist('tiene_ayudante[]')
        costo_ayudante_list = request.form.getlist('costo_ayudante[]')
        
        print(f"DEBUG - Datos del formulario recibidos:")
        print(f"  Form keys: {list(request.form.keys())}")
        print(f"  Categorías: {servicios_categoria}")
        print(f"  Descripciones: {servicios_desc}")
        print(f"  Experiencias: {servicios_exp}")
        print(f"  Valores: {servicios_valor}")
        print(f"  Tiene ayudante: {tiene_ayudante_list}")
        print(f"  Costo ayudante: {costo_ayudante_list}")
        
        # Validar que hay servicios
        if not servicios_desc or not any(desc.strip() for desc in servicios_desc):
            print("DEBUG - No se encontraron servicios válidos")
            flash('Debe agregar al menos un servicio', 'danger')
            return redirect(url_for('index'))
        
        # Guardar cada servicio individualmente
        experiencias_validas = []
        todos_servicios = []
        servicios_guardados = 0
        
        for i in range(len(servicios_desc)):
            if servicios_desc[i].strip():
                categoria = servicios_categoria[i] if i < len(servicios_categoria) else 'Otro'
                exp_valor = float(servicios_exp[i]) if servicios_exp[i] and servicios_exp[i].strip() else 0
                valor_hora = float(servicios_valor[i]) if servicios_valor[i] and servicios_valor[i].strip() else 0
                tiene_ayudante = int(tiene_ayudante_list[i]) if i < len(tiene_ayudante_list) else 0
                costo_ayudante = float(costo_ayudante_list[i]) if i < len(costo_ayudante_list) and costo_ayudante_list[i] and tiene_ayudante == 1 else None
                
                print(f"DEBUG - Guardando servicio {i+1}:")
                print(f"  ID Persona: {id_persona}")
                print(f"  Categoría: '{categoria}'")
                print(f"  Descripción: '{servicios_desc[i].strip()}'")
                print(f"  Experiencia: {exp_valor}")
                print(f"  Valor hora: {valor_hora}")
                print(f"  Tiene ayudante: {tiene_ayudante}")
                print(f"  Costo ayudante: {costo_ayudante}")
                
                try:
                    cursor.execute("""
                        INSERT INTO servicios_persona (id_persona, categoria, descripcion, anios_experiencia, valor_hora, tiene_ayudante, costo_ayudante)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """, (
                        id_persona,
                        categoria,
                        servicios_desc[i].strip(),
                        exp_valor,
                        valor_hora,
                        tiene_ayudante,
                        costo_ayudante
                    ))
                    servicios_guardados += 1
                    print(f"DEBUG - Servicio {i+1} guardado exitosamente")
                except Exception as e:
                    print(f"DEBUG - Error guardando servicio {i+1}: {e}")
                
                experiencias_validas.append(exp_valor)
                todos_servicios.append(f"{categoria}: {servicios_desc[i].strip()}")
        
        print(f"DEBUG - Total servicios guardados: {servicios_guardados} de {len(servicios_desc)}")
        
        # ====== 4. experiencia_persona (resumen general) ======
        # Calcular experiencia máxima y concatenar descripciones
        max_experiencia = max(experiencias_validas) if experiencias_validas else 0
        resumen_servicios = ' | '.join(todos_servicios)
        
        cursor.execute("""
            INSERT INTO experiencia_persona (id_persona, anios_experiencia, descripcion)
            VALUES (%s, %s, %s)
        """, (
            id_persona,
            int(max_experiencia),
            resumen_servicios
        ))
        
        # ====== 5. disponibilidad ======
        cursor.execute("""
            INSERT INTO disponibilidad (id_persona, id_horario, id_dias)
            VALUES (%s, %s, %s)
        """, (
            id_persona,
            request.form['disponibilidad'],
            request.form['disponibilidad_dias']
        ))
        
        # ====== 6. detalles_persona ======
        acepta_terminos = 1 if request.form.get('acepta_terminos') == 'on' else 0
        permisos_ubicacion = 1 if request.form.get('permisos_ubicacion') == 'on' else 0
        
        cursor.execute("""
            INSERT INTO detalles_persona 
            (id_persona, id_servicio_tipo, tareas, antecedentes_pdf, foto_identificacion, 
             acepta_terminos, permisos_ubicacion, recomendaciones, recomendaciones_archivo)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            id_persona,
            request.form['habilidades_tipo'],
            resumen_servicios,  # Usar el resumen de servicios
            antecedentes_filename,
            foto_filename,
            acepta_terminos,
            permisos_ubicacion,
            request.form.get('recomendaciones', ''),
            recomend_filename
        ))
        
        connection.commit()
        print("DEBUG - Transacción completada exitosamente")
        
        # Obtener información para el mensaje de confirmación
        nombre_persona = request.form['nombre_completo']
        fecha_hora = datetime.now().strftime('%d/%m/%Y a las %H:%M:%S')
        
        flash(f'✅ Registro exitoso: {nombre_persona} fue registrado el {fecha_hora}', 'success')
        
    except mysql.connector.Error as err:
        if connection:
            connection.rollback()
        print(f"DEBUG - Error de base de datos: {err}")
        flash(f'Error en la base de datos: {err}', 'danger')
        
    except Exception as e:
        if connection:
            connection.rollback()
        print(f"DEBUG - Error general: {str(e)}")
        flash(f'Error al procesar el formulario: {str(e)}', 'danger')
        
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()
    
    return redirect(url_for('index'))

@app.route('/registros')
@login_required
def ver_registros():
    """Página para ver todos los registros"""
    connection = get_db_connection()
    if not connection:
        return "Error de conexión a la base de datos"
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        # Obtener todos los registros con sus detalles (solo activos)
        cursor.execute("""
            SELECT p.id_persona, p.numero_documento, p.nombre_completo, p.ciudad, p.codigo_dane, 
                   p.fecha_registro, p.registrado_por,
                   td.descripcion as tipo_documento,
                   g.descripcion as genero,
                   tp.telefono,
                   cp.correo,
                   ep.anios_experiencia, ep.descripcion as experiencia_desc,
                   h.descripcion as horario,
                   d.descripcion as dias_disponibles,
                   ht.descripcion as servicio_tipo,
                   dp.foto_identificacion,
                   dp.antecedentes_pdf,
                   dp.recomendaciones_archivo,
                   dp.recomendaciones
            FROM personas p
            LEFT JOIN detalle_parametro td ON p.id_tipo_documento = td.id_detalle
            LEFT JOIN detalle_parametro g ON p.id_genero = g.id_detalle
            LEFT JOIN telefono_persona tp ON p.id_persona = tp.id_persona
            LEFT JOIN correo_persona cp ON p.id_persona = cp.id_persona
            LEFT JOIN experiencia_persona ep ON p.id_persona = ep.id_persona
            LEFT JOIN disponibilidad disp ON p.id_persona = disp.id_persona
            LEFT JOIN detalle_parametro h ON disp.id_horario = h.id_detalle
            LEFT JOIN detalle_parametro d ON disp.id_dias = d.id_detalle
            LEFT JOIN detalles_persona dp ON p.id_persona = dp.id_persona
            LEFT JOIN detalle_parametro ht ON dp.id_servicio_tipo = ht.id_detalle
            WHERE p.estado = 'activo'
            ORDER BY p.fecha_registro DESC
        """)
        
        registros = cursor.fetchall()
        
        # Obtener servicios detallados para cada persona
        for reg in registros:
            cursor.execute("""
                SELECT categoria, descripcion, anios_experiencia, valor_hora, tiene_ayudante, costo_ayudante
                FROM servicios_persona
                WHERE id_persona = %s
                ORDER BY id_servicio
            """, (reg['id_persona'],))
            servicios = cursor.fetchall()
            reg['servicios'] = servicios
            print(f"DEBUG - Persona {reg['nombre_completo']} tiene {len(servicios)} servicios")
        
        # Contar total de registros activos
        cursor.execute("SELECT COUNT(*) as total FROM personas WHERE estado = 'activo'")
        total = cursor.fetchone()['total']
        
        return render_template('registros.html', registros=registros, total=total)
        
    except mysql.connector.Error as err:
        return f"Error al obtener registros: {err}"
    finally:
        if connection and connection.is_connected():
            connection.close()

@app.route('/eliminar/<int:id_persona>', methods=['POST'])
@login_required
def eliminar_registro(id_persona):
    """Marcar registro como eliminado (soft delete) en lugar de eliminarlo permanentemente"""
    connection = get_db_connection()
    if not connection:
        flash('Error de conexión a la base de datos', 'danger')
        return redirect(url_for('ver_registros'))
    
    try:
        cursor = connection.cursor()
        
        # Marcar como eliminado en lugar de borrar
        cursor.execute("""
            UPDATE personas 
            SET estado = 'eliminado', fecha_eliminacion = NOW()
            WHERE id_persona = %s
        """, (id_persona,))
        
        connection.commit()
        
        flash('✅ Registro movido a papelera. Puedes restaurarlo desde "Ver Eliminados".', 'success')
        
    except mysql.connector.Error as err:
        connection.rollback()
        flash(f'Error al eliminar registro: {err}', 'danger')
    finally:
        if connection and connection.is_connected():
            connection.close()
    
    return redirect(url_for('ver_registros'))

@app.route('/eliminar-permanente/<int:id_persona>', methods=['POST'])
@login_required
def eliminar_permanente(id_persona):
    """Eliminar registro permanentemente de la base de datos"""
    connection = get_db_connection()
    if not connection:
        flash('Error de conexión a la base de datos', 'danger')
        return redirect(url_for('ver_eliminados'))
    
    try:
        cursor = connection.cursor()
        
        # Obtener archivos antes de eliminar
        cursor.execute("""
            SELECT foto_identificacion, antecedentes_pdf, recomendaciones_archivo 
            FROM detalles_persona 
            WHERE id_persona = %s
        """, (id_persona,))
        
        archivos = cursor.fetchone()
        
        # Eliminar permanentemente
        cursor.execute("DELETE FROM personas WHERE id_persona = %s", (id_persona,))
        
        # Desactivar temporalmente las foreign keys para reorganizar
        cursor.execute("SET FOREIGN_KEY_CHECKS = 0")
        
        # Obtener todos los IDs actuales ordenados
        cursor.execute("SELECT id_persona FROM personas ORDER BY id_persona")
        ids_actuales = [row[0] for row in cursor.fetchall()]
        
        # Reorganizar IDs en todas las tablas relacionadas
        for nuevo_id, id_actual in enumerate(ids_actuales, start=1):
            if nuevo_id != id_actual:
                cursor.execute("UPDATE personas SET id_persona = %s WHERE id_persona = %s", (nuevo_id, id_actual))
                cursor.execute("UPDATE telefono_persona SET id_persona = %s WHERE id_persona = %s", (nuevo_id, id_actual))
                cursor.execute("UPDATE correo_persona SET id_persona = %s WHERE id_persona = %s", (nuevo_id, id_actual))
                cursor.execute("UPDATE experiencia_persona SET id_persona = %s WHERE id_persona = %s", (nuevo_id, id_actual))
                cursor.execute("UPDATE servicios_persona SET id_persona = %s WHERE id_persona = %s", (nuevo_id, id_actual))
                cursor.execute("UPDATE disponibilidad SET id_persona = %s WHERE id_persona = %s", (nuevo_id, id_actual))
                cursor.execute("UPDATE detalles_persona SET id_persona = %s WHERE id_persona = %s", (nuevo_id, id_actual))
        
        # Resetear AUTO_INCREMENT
        next_id = len(ids_actuales) + 1
        cursor.execute(f"ALTER TABLE personas AUTO_INCREMENT = {next_id}")
        
        # Reactivar foreign keys
        cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
        
        connection.commit()
        
        # Eliminar archivos físicos
        if archivos:
            if archivos[0]:
                foto_path = os.path.join(app.config['UPLOAD_FOLDER'], archivos[0])
                if os.path.exists(foto_path):
                    os.remove(foto_path)
            if archivos[1]:
                antecedentes_path = os.path.join(app.config['UPLOAD_FOLDER'], archivos[1])
                if os.path.exists(antecedentes_path):
                    os.remove(antecedentes_path)
            if archivos[2]:
                recomend_path = os.path.join(app.config['UPLOAD_FOLDER'], archivos[2])
                if os.path.exists(recomend_path):
                    os.remove(recomend_path)
        
        flash(f'✅ Registro eliminado permanentemente. IDs reorganizados. Próximo ID: {next_id}', 'success')
        
    except mysql.connector.Error as err:
        connection.rollback()
        flash(f'Error al eliminar registro: {err}', 'danger')
    finally:
        if connection and connection.is_connected():
            connection.close()
    
    return redirect(url_for('ver_eliminados'))

@app.route('/restaurar/<int:id_persona>', methods=['POST'])
@login_required
def restaurar_registro(id_persona):
    """Restaurar un registro eliminado"""
    connection = get_db_connection()
    if not connection:
        flash('Error de conexión a la base de datos', 'danger')
        return redirect(url_for('ver_eliminados'))
    
    try:
        cursor = connection.cursor()
        
        # Restaurar registro
        cursor.execute("""
            UPDATE personas 
            SET estado = 'activo', fecha_eliminacion = NULL
            WHERE id_persona = %s
        """, (id_persona,))
        
        connection.commit()
        
        flash('✅ Registro restaurado exitosamente.', 'success')
        
    except mysql.connector.Error as err:
        connection.rollback()
        flash(f'Error al restaurar registro: {err}', 'danger')
    finally:
        if connection and connection.is_connected():
            connection.close()
    
    return redirect(url_for('ver_eliminados'))

@app.route('/eliminados')
@login_required
def ver_eliminados():
    """Página para ver registros eliminados (papelera)"""
    connection = get_db_connection()
    if not connection:
        return "Error de conexión a la base de datos"
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        # Obtener registros eliminados
        cursor.execute("""
            SELECT p.id_persona, p.numero_documento, p.nombre_completo, p.ciudad, 
                   p.fecha_eliminacion,
                   td.descripcion as tipo_documento,
                   tp.telefono
            FROM personas p
            LEFT JOIN detalle_parametro td ON p.id_tipo_documento = td.id_detalle
            LEFT JOIN telefono_persona tp ON p.id_persona = tp.id_persona
            WHERE p.estado = 'eliminado'
            ORDER BY p.fecha_eliminacion DESC
        """)
        
        registros = cursor.fetchall()
        
        # Contar total
        cursor.execute("SELECT COUNT(*) as total FROM personas WHERE estado = 'eliminado'")
        total = cursor.fetchone()['total']
        
        return render_template('eliminados.html', registros=registros, total=total)
        
    except mysql.connector.Error as err:
        return f"Error al obtener registros: {err}"
    finally:
        if connection and connection.is_connected():
            connection.close()

@app.route('/debug-db')
def debug_db():
    """Ruta temporal para debuggear la base de datos"""
    connection = get_db_connection()
    if not connection:
        return "Error de conexión"
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        # Verificar si existe la tabla servicios_persona
        cursor.execute("SHOW TABLES LIKE 'servicios_persona'")
        tabla_existe = cursor.fetchone()
        
        resultado = f"<h2>Debug Base de Datos</h2>"
        resultado += f"<p>Tabla servicios_persona existe: {'SÍ' if tabla_existe else 'NO'}</p>"
        
        if tabla_existe:
            # Mostrar estructura
            cursor.execute("DESCRIBE servicios_persona")
            estructura = cursor.fetchall()
            resultado += "<h3>Estructura de servicios_persona:</h3><ul>"
            for col in estructura:
                resultado += f"<li>{col['Field']} - {col['Type']}</li>"
            resultado += "</ul>"
            
            # Contar registros
            cursor.execute("SELECT COUNT(*) as total FROM servicios_persona")
            total = cursor.fetchone()['total']
            resultado += f"<p>Total servicios: {total}</p>"
            
            # Mostrar algunos registros
            if total > 0:
                cursor.execute("SELECT * FROM servicios_persona LIMIT 10")
                registros = cursor.fetchall()
                resultado += "<h3>Registros:</h3><table border='1'>"
                resultado += "<tr><th>ID</th><th>Persona</th><th>Descripción</th><th>Experiencia</th><th>Valor/Hora</th></tr>"
                for reg in registros:
                    resultado += f"<tr><td>{reg['id_servicio']}</td><td>{reg['id_persona']}</td><td>{reg['descripcion'][:50]}...</td><td>{reg['anios_experiencia']}</td><td>{reg['valor_hora']}</td></tr>"
                resultado += "</table>"
        
        # Verificar personas
        cursor.execute("SELECT COUNT(*) as total FROM personas")
        total_personas = cursor.fetchone()['total']
        resultado += f"<p>Total personas: {total_personas}</p>"
        
        return resultado
        
    except Exception as e:
        return f"Error: {e}"
    finally:
        if connection and connection.is_connected():
            connection.close()

@app.route('/actualizar-db')
def actualizar_db():
    """Ruta temporal para actualizar la base de datos"""
    try:
        init_database()
        return "Base de datos actualizada exitosamente. <a href='/registros'>Ver registros</a>"
    except Exception as e:
        return f"Error al actualizar base de datos: {e}"

@app.route('/buscar-registros')
@login_required
def buscar_registros():
    """API para búsqueda y filtrado de registros"""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Error de conexión'}), 500
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        # Obtener parámetros de búsqueda
        search = request.args.get('search', '').strip()
        categoria = request.args.get('categoria', '').strip()
        tarifa_min = request.args.get('tarifa_min', '').strip()
        tarifa_max = request.args.get('tarifa_max', '').strip()
        
        # Construir query base
        query = """
            SELECT DISTINCT p.id_persona, p.numero_documento, p.nombre_completo, p.ciudad, p.codigo_dane, 
                   p.fecha_registro, p.registrado_por,
                   td.descripcion as tipo_documento,
                   g.descripcion as genero,
                   tp.telefono,
                   cp.correo,
                   ep.anios_experiencia, ep.descripcion as experiencia_desc,
                   h.descripcion as horario,
                   d.descripcion as dias_disponibles,
                   ht.descripcion as servicio_tipo,
                   dp.foto_identificacion,
                   dp.antecedentes_pdf,
                   dp.recomendaciones_archivo,
                   dp.recomendaciones
            FROM personas p
            LEFT JOIN detalle_parametro td ON p.id_tipo_documento = td.id_detalle
            LEFT JOIN detalle_parametro g ON p.id_genero = g.id_detalle
            LEFT JOIN telefono_persona tp ON p.id_persona = tp.id_persona
            LEFT JOIN correo_persona cp ON p.id_persona = cp.id_persona
            LEFT JOIN experiencia_persona ep ON p.id_persona = ep.id_persona
            LEFT JOIN disponibilidad disp ON p.id_persona = disp.id_persona
            LEFT JOIN detalle_parametro h ON disp.id_horario = h.id_detalle
            LEFT JOIN detalle_parametro d ON disp.id_dias = d.id_detalle
            LEFT JOIN detalles_persona dp ON p.id_persona = dp.id_persona
            LEFT JOIN detalle_parametro ht ON dp.id_servicio_tipo = ht.id_detalle
            LEFT JOIN servicios_persona sp ON p.id_persona = sp.id_persona
            WHERE 1=1
        """
        
        params = []
        
        # Filtro de búsqueda por texto
        if search:
            query += """ AND (p.nombre_completo LIKE %s 
                         OR p.numero_documento LIKE %s 
                         OR p.ciudad LIKE %s)"""
            search_param = f"%{search}%"
            params.extend([search_param, search_param, search_param])
        
        # Filtro por categoría
        if categoria:
            query += " AND sp.categoria = %s"
            params.append(categoria)
        
        # Filtro por rango de tarifas
        if tarifa_min:
            query += " AND sp.valor_hora >= %s"
            params.append(float(tarifa_min))
        
        if tarifa_max:
            query += " AND sp.valor_hora <= %s"
            params.append(float(tarifa_max))
        
        query += " ORDER BY p.fecha_registro DESC"
        
        cursor.execute(query, params)
        registros = cursor.fetchall()
        
        # Obtener servicios para cada persona
        for reg in registros:
            cursor.execute("""
                SELECT categoria, descripcion, anios_experiencia, valor_hora, tiene_ayudante, costo_ayudante
                FROM servicios_persona
                WHERE id_persona = %s
                ORDER BY id_servicio
            """, (reg['id_persona'],))
            servicios = cursor.fetchall()
            reg['servicios'] = servicios
        
        return jsonify({
            'success': True,
            'registros': registros,
            'total': len(registros)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if connection and connection.is_connected():
            connection.close()

@app.route('/exportar-excel')
@login_required
def exportar_excel():
    """Exportar todos los registros a Excel"""
    connection = get_db_connection()
    if not connection:
        flash('Error de conexión a la base de datos', 'danger')
        return redirect(url_for('ver_registros'))
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        # Obtener todos los registros
        cursor.execute("""
            SELECT p.id_persona, p.numero_documento, p.nombre_completo, p.ciudad, p.codigo_dane, 
                   p.fecha_registro, p.registrado_por,
                   td.descripcion as tipo_documento,
                   g.descripcion as genero,
                   tp.telefono,
                   cp.correo,
                   ep.anios_experiencia,
                   h.descripcion as horario,
                   d.descripcion as dias_disponibles,
                   ht.descripcion as servicio_tipo
            FROM personas p
            LEFT JOIN detalle_parametro td ON p.id_tipo_documento = td.id_detalle
            LEFT JOIN detalle_parametro g ON p.id_genero = g.id_detalle
            LEFT JOIN telefono_persona tp ON p.id_persona = tp.id_persona
            LEFT JOIN correo_persona cp ON p.id_persona = cp.id_persona
            LEFT JOIN experiencia_persona ep ON p.id_persona = ep.id_persona
            LEFT JOIN disponibilidad disp ON p.id_persona = disp.id_persona
            LEFT JOIN detalle_parametro h ON disp.id_horario = h.id_detalle
            LEFT JOIN detalle_parametro d ON disp.id_dias = d.id_detalle
            LEFT JOIN detalles_persona dp ON p.id_persona = dp.id_persona
            LEFT JOIN detalle_parametro ht ON dp.id_servicio_tipo = ht.id_detalle
            ORDER BY p.fecha_registro DESC
        """)
        
        registros = cursor.fetchall()
        
        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Profesionales"
        
        # Estilos
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Encabezados
        headers = [
            'ID', 'Tipo Doc', 'Número Doc', 'Nombre Completo', 'Género', 
            'Ciudad', 'Código DANE', 'Teléfono', 'Correo', 'Experiencia (años)',
            'Tipo Servicios', 'Horario', 'Días', 'Servicios Ofrecidos', 
            'Tarifas (COP/h)', 'Registrado Por', 'Fecha Registro'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        for row_idx, reg in enumerate(registros, 2):
            # Obtener servicios
            cursor.execute("""
                SELECT categoria, descripcion, valor_hora, tiene_ayudante, costo_ayudante
                FROM servicios_persona
                WHERE id_persona = %s
            """, (reg['id_persona'],))
            servicios = cursor.fetchall()
            
            servicios_text = []
            tarifas_text = []
            for serv in servicios:
                servicios_text.append(f"{serv['categoria']}: {serv['descripcion'][:50]}")
                tarifa = f"${serv['valor_hora']:,.0f}"
                if serv['tiene_ayudante'] == 1:
                    tarifa += f" (+Ayudante: ${serv['costo_ayudante']:,.0f})"
                tarifas_text.append(tarifa)
            
            ws.cell(row=row_idx, column=1, value=reg['id_persona'])
            ws.cell(row=row_idx, column=2, value=reg['tipo_documento'])
            ws.cell(row=row_idx, column=3, value=reg['numero_documento'])
            ws.cell(row=row_idx, column=4, value=reg['nombre_completo'])
            ws.cell(row=row_idx, column=5, value=reg['genero'])
            ws.cell(row=row_idx, column=6, value=reg['ciudad'])
            ws.cell(row=row_idx, column=7, value=reg['codigo_dane'])
            ws.cell(row=row_idx, column=8, value=reg['telefono'])
            ws.cell(row=row_idx, column=9, value=reg['correo'] or 'N/A')
            ws.cell(row=row_idx, column=10, value=reg['anios_experiencia'])
            ws.cell(row=row_idx, column=11, value=reg['servicio_tipo'])
            ws.cell(row=row_idx, column=12, value=reg['horario'])
            ws.cell(row=row_idx, column=13, value=reg['dias_disponibles'])
            ws.cell(row=row_idx, column=14, value=' | '.join(servicios_text))
            ws.cell(row=row_idx, column=15, value=' | '.join(tarifas_text))
            ws.cell(row=row_idx, column=16, value=reg['registrado_por'])
            ws.cell(row=row_idx, column=17, value=reg['fecha_registro'].strftime('%d/%m/%Y %H:%M') if reg['fecha_registro'] else 'N/A')
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo con fecha
        filename = f"profesionales_talenthub_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error al exportar: {str(e)}', 'danger')
        return redirect(url_for('ver_registros'))
    finally:
        if connection and connection.is_connected():
            connection.close()

@app.route('/estadisticas')
@login_required
def estadisticas():
    """Obtener estadísticas del sistema"""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Error de conexión'}), 500
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        # Total de profesionales
        cursor.execute("SELECT COUNT(*) as total FROM personas")
        total_profesionales = cursor.fetchone()['total']
        
        # Profesionales por categoría
        cursor.execute("""
            SELECT categoria, COUNT(*) as total
            FROM servicios_persona
            GROUP BY categoria
            ORDER BY total DESC
        """)
        por_categoria = cursor.fetchall()
        
        # Promedio de tarifas por categoría
        cursor.execute("""
            SELECT categoria, 
                   AVG(valor_hora) as promedio,
                   MIN(valor_hora) as minimo,
                   MAX(valor_hora) as maximo
            FROM servicios_persona
            WHERE valor_hora > 0
            GROUP BY categoria
            ORDER BY promedio DESC
        """)
        tarifas_categoria = cursor.fetchall()
        
        # Distribución geográfica
        cursor.execute("""
            SELECT ciudad, COUNT(*) as total
            FROM personas
            GROUP BY ciudad
            ORDER BY total DESC
            LIMIT 10
        """)
        por_ciudad = cursor.fetchall()
        
        # Profesionales con más experiencia
        cursor.execute("""
            SELECT p.nombre_completo, p.ciudad, ep.anios_experiencia
            FROM personas p
            JOIN experiencia_persona ep ON p.id_persona = ep.id_persona
            ORDER BY ep.anios_experiencia DESC
            LIMIT 5
        """)
        mas_experimentados = cursor.fetchall()
        
        return jsonify({
            'total_profesionales': total_profesionales,
            'por_categoria': por_categoria,
            'tarifas_categoria': tarifas_categoria,
            'por_ciudad': por_ciudad,
            'mas_experimentados': mas_experimentados
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if connection and connection.is_connected():
            connection.close()

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    """Página de login para administrador"""
    if request.method == 'POST':
        password = request.form.get('password', '')
        
        # Verificar contraseña con bcrypt
        if bcrypt.checkpw(password.encode('utf-8'), ADMIN_PASSWORD_HASH):
            session['admin_logged_in'] = True
            session['last_activity'] = datetime.now().isoformat()
            session.permanent = True
            flash('Inicio de sesión exitoso', 'success')
            return redirect(url_for('ver_registros'))
        else:
            flash('Contraseña incorrecta', 'danger')
    
    return render_template('admin_login.html')

@app.route('/admin/logout')
def admin_logout():
    """Cerrar sesión de administrador"""
    session.pop('admin_logged_in', None)
    flash('Sesión cerrada exitosamente', 'info')
    return redirect(url_for('index'))

@app.route('/reorganizar-ids')
@login_required
def reorganizar_ids():
    """Reorganizar todos los IDs de la tabla personas para mantener secuencia continua"""
    connection = get_db_connection()
    if not connection:
        flash('Error de conexión a la base de datos', 'danger')
        return redirect(url_for('ver_registros'))
    
    try:
        cursor = connection.cursor()
        
        # Desactivar temporalmente las foreign keys
        cursor.execute("SET FOREIGN_KEY_CHECKS = 0")
        
        # Obtener todos los registros ordenados por id_persona
        cursor.execute("SELECT id_persona FROM personas ORDER BY id_persona")
        ids_actuales = [row[0] for row in cursor.fetchall()]
        
        # Reorganizar IDs
        for nuevo_id, id_actual in enumerate(ids_actuales, start=1):
            if nuevo_id != id_actual:
                # Actualizar en todas las tablas relacionadas
                cursor.execute("UPDATE personas SET id_persona = %s WHERE id_persona = %s", (nuevo_id, id_actual))
                cursor.execute("UPDATE telefono_persona SET id_persona = %s WHERE id_persona = %s", (nuevo_id, id_actual))
                cursor.execute("UPDATE correo_persona SET id_persona = %s WHERE id_persona = %s", (nuevo_id, id_actual))
                cursor.execute("UPDATE experiencia_persona SET id_persona = %s WHERE id_persona = %s", (nuevo_id, id_actual))
                cursor.execute("UPDATE servicios_persona SET id_persona = %s WHERE id_persona = %s", (nuevo_id, id_actual))
                cursor.execute("UPDATE disponibilidad SET id_persona = %s WHERE id_persona = %s", (nuevo_id, id_actual))
                cursor.execute("UPDATE detalles_persona SET id_persona = %s WHERE id_persona = %s", (nuevo_id, id_actual))
        
        # Resetear AUTO_INCREMENT
        next_id = len(ids_actuales) + 1
        cursor.execute(f"ALTER TABLE personas AUTO_INCREMENT = {next_id}")
        
        # Reactivar foreign keys
        cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
        
        connection.commit()
        
        flash(f'✅ IDs reorganizados exitosamente. Próximo ID: {next_id}', 'success')
        
    except mysql.connector.Error as err:
        connection.rollback()
        flash(f'Error al reorganizar IDs: {err}', 'danger')
    finally:
        if connection and connection.is_connected():
            connection.close()
    
    return redirect(url_for('ver_registros'))

if __name__ == "__main__":
    # Inicializar la base de datos al iniciar la aplicación
    init_database()
    
    # Verificar que la carpeta de uploads existe
    if not os.path.exists(app.config['UPLOAD_FOLDER']):
        os.makedirs(app.config['UPLOAD_FOLDER'])
    
    app.run(debug=True, port=5000)
