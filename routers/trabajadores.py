from fastapi import APIRouter, Form, Request, UploadFile, File, HTTPException, Response
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
import mysql.connector
from datetime import datetime, timedelta
import bcrypt
import auth  # Módulo de autenticación
from config import DB_CONFIG

router = APIRouter(prefix="/trabajador", tags=["trabajadores"])
templates = Jinja2Templates(directory="templates")

# Conexión a BD
def conectar_bd():
    return mysql.connector.connect(**DB_CONFIG)

# ============================================
# TEMPLATES - Aquí conectarás tu formulario de hoja de vida
# ============================================

@router.get("/crear_password", response_class=HTMLResponse)
def mostrar_crear_password(request: Request):
    """Pantalla para crear contraseña después del registro"""
    return templates.TemplateResponse("trabajadores/crear_password.html", {"request": request})

@router.get("/panel", response_class=HTMLResponse)
def mostrar_panel_trabajador(request: Request):
    """Panel principal del trabajador"""
    return templates.TemplateResponse("trabajadores/panel.html", {"request": request})

@router.get("/mi-perfil")
def obtener_mi_perfil(request: Request):
    """Obtiene datos del trabajador autenticado"""
    token = request.cookies.get("session_token")
    if not token:
        return JSONResponse({"error": "No autenticado"}, status_code=401)
    sesion = auth.verificar_sesion(token)
    if not sesion or sesion['tipo_usuario'] != 'trabajador':
        return JSONResponse({"error": "Sesión inválida"}, status_code=401)
    conexion = conectar_bd()
    cursor = conexion.cursor(dictionary=True)
    cursor.execute("""
        SELECT p.id_persona, p.nombre_completo, p.numero_documento, p.ciudad, p.departamento,
               tp.telefono, cp.correo, cp.verificado AS email_verificado
        FROM personas p
        LEFT JOIN telefono_persona tp ON p.id_persona = tp.id_persona
        LEFT JOIN correo_persona cp ON p.id_persona = cp.id_persona
        WHERE p.id_persona = %s
    """, (sesion['id_usuario'],))
    trabajador = cursor.fetchone()
    cursor.close()
    conexion.close()
    if not trabajador:
        return JSONResponse({"error": "No encontrado"}, status_code=404)
    # Serializar
    for k, v in trabajador.items():
        if v is None:
            trabajador[k] = ''
    return JSONResponse(trabajador)

@router.get("/solicitudes_pendientes_panel", response_class=HTMLResponse)
def mostrar_mis_solicitudes_panel(request: Request):
    """Panel de solicitudes del trabajador"""
    return templates.TemplateResponse("trabajadores/mis_solicitudes.html", {"request": request})

@router.get("/mis_solicitudes_api")
def listar_mis_solicitudes(id_trabajador: int = None):
    """Solicitudes asignadas o pendientes para un trabajador, filtradas por sus categorías"""
    conexion = conectar_bd()
    if not conexion:
        return JSONResponse({"error": "Error de conexión", "solicitudes": []}, status_code=500)
    try:
        cursor = conexion.cursor(dictionary=True)
        if id_trabajador:
            # Obtener las categorías del trabajador (texto)
            cursor.execute("""
                SELECT DISTINCT categoria FROM servicios_persona WHERE id_persona = %s
            """, (id_trabajador,))
            cats = [r['categoria'] for r in cursor.fetchall()]

            # Buscar solicitudes propias + pendientes que coincidan con sus categorías
            if cats:
                placeholders = ','.join(['%s'] * len(cats))
                cursor.execute(f"""
                    SELECT s.id_solicitud, s.id_cliente, s.titulo, s.descripcion, s.estado,
                           s.ciudad, s.departamento, s.direccion_servicio, s.fecha_solicitud,
                           s.metodo_pago,
                           COALESCE(cat.nombre_categoria, s.titulo) as nombre_categoria,
                           c.nombre_completo as nombre_cliente,
                           tc.telefono as telefono_cliente
                    FROM solicitudes_servicio s
                    LEFT JOIN categorias_servicio cat ON s.id_categoria = cat.id_categoria
                    LEFT JOIN clientes c ON s.id_cliente = c.id_cliente
                    LEFT JOIN telefono_cliente tc ON c.id_cliente = tc.id_cliente
                    WHERE s.id_trabajador = %s
                       OR (s.estado = 'pendiente' AND cat.nombre_categoria IN ({placeholders}))
                    ORDER BY s.fecha_solicitud DESC
                    LIMIT 50
                """, (id_trabajador, *cats))
            else:
                cursor.execute("""
                    SELECT s.id_solicitud, s.id_cliente, s.titulo, s.descripcion, s.estado,
                           s.ciudad, s.departamento, s.direccion_servicio, s.fecha_solicitud,
                           s.metodo_pago,
                           COALESCE(cat.nombre_categoria, s.titulo) as nombre_categoria,
                           c.nombre_completo as nombre_cliente,
                           tc.telefono as telefono_cliente
                    FROM solicitudes_servicio s
                    LEFT JOIN categorias_servicio cat ON s.id_categoria = cat.id_categoria
                    LEFT JOIN clientes c ON s.id_cliente = c.id_cliente
                    LEFT JOIN telefono_cliente tc ON c.id_cliente = tc.id_cliente
                    WHERE s.id_trabajador = %s
                    ORDER BY s.fecha_solicitud DESC
                    LIMIT 50
                """, (id_trabajador,))
        else:
            cursor.execute("""
                SELECT s.id_solicitud, s.id_cliente, s.titulo, s.descripcion, s.estado,
                       s.ciudad, s.departamento, s.direccion_servicio, s.fecha_solicitud,
                       COALESCE(cat.nombre_categoria, s.titulo) as nombre_categoria,
                       c.nombre_completo as nombre_cliente,
                       tc.telefono as telefono_cliente
                FROM solicitudes_servicio s
                LEFT JOIN categorias_servicio cat ON s.id_categoria = cat.id_categoria
                LEFT JOIN clientes c ON s.id_cliente = c.id_cliente
                LEFT JOIN telefono_cliente tc ON c.id_cliente = tc.id_cliente
                WHERE s.estado = 'pendiente'
                ORDER BY s.fecha_solicitud DESC
                LIMIT 20
            """)
        solicitudes = cursor.fetchall()
        for s in solicitudes:
            for k, v in s.items():
                if hasattr(v, 'isoformat'):
                    s[k] = v.strftime('%Y-%m-%d %H:%M')
                elif v is None:
                    s[k] = ''
        return JSONResponse({"solicitudes": solicitudes})
    except Exception as e:
        return JSONResponse({"error": str(e), "solicitudes": []}, status_code=500)
    finally:
        if conexion and conexion.is_connected():
            conexion.close()


@router.post("/solicitud/actualizar")
def actualizar_estado_solicitud(
    id_solicitud: int = Form(...),
    estado: str = Form(...),
    id_trabajador: int = Form(None)
):
    """Permite al trabajador aceptar, rechazar, iniciar o completar una solicitud"""
    estados_validos = ['aceptada', 'cancelada', 'completada', 'en_proceso']
    if estado not in estados_validos:
        return JSONResponse({"error": "Estado no válido"}, status_code=400)

    conexion = conectar_bd()
    if not conexion:
        return JSONResponse({"error": "Error de conexión"}, status_code=500)
    try:
        cursor = conexion.cursor()

        if estado == 'aceptada' and id_trabajador:
            # Race condition fix: solo actualizar si todavía está pendiente
            cursor.execute("""
                UPDATE solicitudes_servicio
                SET estado = %s, id_trabajador = %s, fecha_aceptacion = NOW()
                WHERE id_solicitud = %s AND estado = 'pendiente'
            """, (estado, id_trabajador, id_solicitud))
            if cursor.rowcount == 0:
                conexion.rollback()
                return JSONResponse(
                    {"error": "Esta solicitud ya fue aceptada por otro trabajador"},
                    status_code=409
                )

        elif estado == 'cancelada':
            # Verificar el estado actual de la solicitud
            cursor_check = conexion.cursor()
            cursor_check.execute("""
                SELECT estado, id_trabajador FROM solicitudes_servicio
                WHERE id_solicitud = %s
            """, (id_solicitud,))
            sol = cursor_check.fetchone()
            cursor_check.close()

            if sol and sol[0] in ('pendiente', 'aceptada') and sol[1]:
                # Tenía trabajador asignado → liberar: volver a pendiente sin trabajador
                # El cliente no tiene que repetir el proceso, otros trabajadores pueden tomar la solicitud
                cursor.execute("""
                    UPDATE solicitudes_servicio
                    SET estado = 'pendiente', id_trabajador = NULL, fecha_aceptacion = NULL
                    WHERE id_solicitud = %s
                """, (id_solicitud,))
                conexion.commit()
                return JSONResponse({"mensaje": "Solicitud liberada — otros profesionales pueden tomarla"})
            else:
                # Sin trabajador asignado (cancelación del cliente) o en proceso → cancelar normal
                cursor.execute("""
                    UPDATE solicitudes_servicio
                    SET estado = 'cancelada'
                    WHERE id_solicitud = %s
                """, (id_solicitud,))

        elif estado == 'en_proceso':
            # Registrar fecha de inicio del servicio
            cursor.execute("""
                UPDATE solicitudes_servicio
                SET estado = %s, fecha_inicio = NOW()
                WHERE id_solicitud = %s AND estado = 'aceptada'
            """, (estado, id_solicitud))

        elif estado == 'completada':
            # Registrar fecha de finalización
            cursor.execute("""
                UPDATE solicitudes_servicio
                SET estado = %s, fecha_finalizacion = NOW()
                WHERE id_solicitud = %s AND estado = 'en_proceso'
            """, (estado, id_solicitud))

        else:
            cursor.execute("""
                UPDATE solicitudes_servicio
                SET estado = %s
                WHERE id_solicitud = %s
            """, (estado, id_solicitud))

        conexion.commit()
        return JSONResponse({"mensaje": f"Solicitud actualizada a '{estado}'"})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        if conexion and conexion.is_connected():
            conexion.close()

@router.get("/perfil_completo")
def obtener_perfil_completo(id: int = None):
    """Obtiene el perfil completo del trabajador"""
    if not id:
        return JSONResponse({"error": "ID requerido"}, status_code=400)
    conexion = conectar_bd()
    try:
        cursor = conexion.cursor(dictionary=True)
        cursor.execute("""
            SELECT p.id_persona, p.nombre_completo, p.numero_documento,
                   p.ciudad, p.departamento, p.nacionalidad, p.fecha_nacimiento,
                   tp.telefono, cp.correo
            FROM personas p
            LEFT JOIN telefono_persona tp ON p.id_persona = tp.id_persona
            LEFT JOIN correo_persona cp ON p.id_persona = cp.id_persona
            WHERE p.id_persona = %s
        """, (id,))
        perfil = cursor.fetchone()
        if not perfil:
            return JSONResponse({"error": "No encontrado"}, status_code=404)
        for k, v in perfil.items():
            if v is None: perfil[k] = ''
            elif hasattr(v, 'isoformat'): perfil[k] = str(v)
        cursor.execute("SELECT categoria, descripcion, valor_hora, anios_experiencia FROM servicios_persona WHERE id_persona = %s", (id,))
        servicios = cursor.fetchall()
        for s in servicios:
            for k, v in s.items():
                if hasattr(v, '__float__'): s[k] = float(v)
                elif v is None: s[k] = ''
        perfil['servicios'] = servicios
        return JSONResponse(perfil)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        if conexion and conexion.is_connected():
            conexion.close()

@router.get("/mi_perfil_page", response_class=HTMLResponse)
def mostrar_mi_perfil(request: Request):
    return templates.TemplateResponse("trabajadores/mi_perfil.html", {"request": request})

@router.post("/perfil/editar")
def editar_perfil_trabajador(
    id_persona:    int = Form(...),
    ciudad:        str = Form(None),
    departamento:  str = Form(None),
    nacionalidad:  str = Form(None),
    telefono:      str = Form(None),
    correo:        str = Form(None)
):
    """Permite al trabajador actualizar sus datos básicos"""
    conexion = conectar_bd()
    try:
        cursor = conexion.cursor()

        # Actualizar persona (solo si vienen con valor)
        if ciudad or departamento or nacionalidad:
            cursor.execute("""
                UPDATE personas
                SET ciudad = COALESCE(NULLIF(%s,''), ciudad),
                    departamento = COALESCE(NULLIF(%s,''), departamento),
                    nacionalidad = COALESCE(NULLIF(%s,''), nacionalidad)
                WHERE id_persona = %s
            """, (ciudad or '', departamento or '', nacionalidad or '', id_persona))

        # Actualizar teléfono
        if telefono:
            cursor.execute("SELECT 1 FROM telefono_persona WHERE id_persona = %s LIMIT 1", (id_persona,))
            if cursor.fetchone():
                cursor.execute("UPDATE telefono_persona SET telefono = %s WHERE id_persona = %s", (telefono, id_persona))
            else:
                cursor.execute("INSERT INTO telefono_persona (id_persona, telefono) VALUES (%s, %s)", (id_persona, telefono))

        # Actualizar correo
        if correo:
            cursor.execute("SELECT 1 FROM correo_persona WHERE id_persona = %s LIMIT 1", (id_persona,))
            if cursor.fetchone():
                cursor.execute("UPDATE correo_persona SET correo = %s WHERE id_persona = %s", (correo, id_persona))
            else:
                cursor.execute("INSERT INTO correo_persona (id_persona, correo) VALUES (%s, %s)", (id_persona, correo))

        conexion.commit()
        return JSONResponse({"mensaje": "Perfil actualizado correctamente"})
    except Exception as e:
        conexion.rollback()
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        if conexion and conexion.is_connected():
            conexion.close()

@router.post("/baja")
def dar_de_baja_trabajador(id_persona: int = Form(...)):
    """Desactiva la cuenta del trabajador (soft delete)"""
    conexion = conectar_bd()
    try:
        cursor = conexion.cursor()
        cursor.execute("UPDATE personas SET estado = 'inactivo' WHERE id_persona = %s", (id_persona,))
        # Desactivar disponibilidad
        cursor.execute("UPDATE disponibilidad SET disponible = 0 WHERE id_persona = %s", (id_persona,))
        conexion.commit()
        return JSONResponse({"mensaje": "Cuenta desactivada"})
    except Exception as e:
        conexion.rollback()
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        if conexion and conexion.is_connected():
            conexion.close()

@router.get("/perfil/{id_persona}", response_class=HTMLResponse)
def perfil_publico(request: Request, id_persona: int):
    """Perfil público del trabajador — SSR con todos los datos"""
    from datetime import timedelta
    from fastapi.responses import HTMLResponse as _HTML
    from fastapi import status

    conexion = conectar_bd()
    try:
        cursor = conexion.cursor(dictionary=True)

        # Datos base del trabajador
        cursor.execute("""
            SELECT p.id_persona, p.nombre_completo, p.ciudad, p.departamento,
                   p.estado,
                   tp.telefono,
                   dp.foto_identificacion,
                   (dp.foto_identificacion_data IS NOT NULL
                    AND LENGTH(dp.foto_identificacion_data) > 0) AS tiene_foto
            FROM personas p
            LEFT JOIN telefono_persona tp ON p.id_persona = tp.id_persona
            LEFT JOIN detalles_persona dp ON p.id_persona = dp.id_persona
            WHERE p.id_persona = %s
              AND (p.estado = 'activo' OR p.estado IS NULL)
            LIMIT 1
        """, (id_persona,))
        perfil = cursor.fetchone()

        if not perfil:
            raise HTTPException(status_code=404, detail="Trabajador no encontrado")

        # Serializar nulos
        for k, v in perfil.items():
            if v is None:
                perfil[k] = ''

        # Servicios / categorías
        cursor.execute("""
            SELECT categoria, descripcion, anios_experiencia, valor_hora,
                   tiene_ayudante, costo_ayudante
            FROM servicios_persona
            WHERE id_persona = %s
            ORDER BY valor_hora DESC
        """, (id_persona,))
        servicios = cursor.fetchall()
        for s in servicios:
            for k, v in s.items():
                if v is None:
                    s[k] = ''
                elif isinstance(v, float) or (hasattr(v, '__class__') and v.__class__.__name__ == 'Decimal'):
                    s[k] = float(v)

        # Calificación promedio y trabajos completados
        cursor.execute("""
            SELECT ROUND(AVG(puntuacion), 1) AS promedio, COUNT(*) AS total
            FROM calificaciones
            WHERE id_trabajador = %s
        """, (id_persona,))
        cal = cursor.fetchone()
        calificacion   = float(cal['promedio']) if cal and cal['promedio'] else 0.0
        total_resenas  = int(cal['total'])       if cal else 0

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM solicitudes_servicio
            WHERE id_trabajador = %s AND estado = 'completada'
        """, (id_persona,))
        row = cursor.fetchone()
        trabajos_completados = int(row['total']) if row else 0

        # Años de experiencia (máximo entre servicios)
        max_exp = 0
        for s in servicios:
            try:
                exp = float(s.get('anios_experiencia') or 0)
                if exp > max_exp:
                    max_exp = exp
            except Exception:
                pass
        anios_experiencia = int(max_exp)

        # Disponibilidad
        horarios = {7: '8am – 12pm', 8: '2pm – 6pm', 9: '6pm – 10pm', 10: '24 horas'}
        dias_map = {11: 'Entre semana', 12: 'Toda la semana'}
        cursor.execute("""
            SELECT id_horario, id_dias, disponible
            FROM disponibilidad
            WHERE id_persona = %s
            LIMIT 1
        """, (id_persona,))
        disp = cursor.fetchone()
        horario_texto  = horarios.get(disp['id_horario'], '') if disp else ''
        dias_texto     = dias_map.get(disp['id_dias'], '')    if disp else ''
        esta_disponible = bool(disp and disp.get('disponible')) if disp else False

        # Reseñas (últimas 8)
        cursor.execute("""
            SELECT cal.puntuacion, cal.comentario, cal.fecha_calificacion,
                   COALESCE(c.nombre_completo, 'Cliente') AS nombre_cliente
            FROM calificaciones cal
            LEFT JOIN clientes c ON cal.id_cliente = c.id_cliente
            WHERE cal.id_trabajador = %s
              AND cal.comentario IS NOT NULL
              AND cal.comentario != ''
            ORDER BY cal.fecha_calificacion DESC
            LIMIT 8
        """, (id_persona,))
        resenas = cursor.fetchall()
        for r in resenas:
            if r.get('fecha_calificacion') and hasattr(r['fecha_calificacion'], 'isoformat'):
                r['fecha_calificacion'] = (r['fecha_calificacion'] - timedelta(hours=5)).strftime('%d/%m/%Y')
            r['puntuacion'] = int(r['puntuacion']) if r.get('puntuacion') else 0
            if not r.get('comentario'):
                r['comentario'] = ''

        # Tarifa principal (mayor valor_hora)
        tarifa_principal = 0.0
        tiene_ayudante   = False
        costo_ayudante   = 0.0
        for s in servicios:
            vh = float(s.get('valor_hora') or 0)
            if vh > tarifa_principal:
                tarifa_principal = vh
            if s.get('tiene_ayudante') == 1:
                tiene_ayudante = True
                ca = float(s.get('costo_ayudante') or 0)
                if ca > costo_ayudante:
                    costo_ayudante = ca

        # Verificado: tiene documentos cargados
        verificado = bool(perfil.get('tiene_foto'))

        # Iniciales del nombre
        partes = str(perfil.get('nombre_completo', '')).split()
        iniciales = ''.join(p[0].upper() for p in partes[:2]) if partes else '?'

        # Miembro desde
        cursor.execute("SELECT fecha_registro FROM personas WHERE id_persona = %s", (id_persona,))
        row_fecha = cursor.fetchone()
        if row_fecha and row_fecha.get('fecha_registro'):
            fecha_reg = row_fecha['fecha_registro']
            meses_es = ['','Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
            miembro_desde = f"{meses_es[fecha_reg.month]} {fecha_reg.year}"
        else:
            miembro_desde = ""

        # Últimos trabajos completados (historial público)
        cursor.execute("""
            SELECT s.titulo, cat.nombre_categoria, s.fecha_finalizacion, s.ciudad
            FROM solicitudes_servicio s
            LEFT JOIN categorias_servicio cat ON s.id_categoria = cat.id_categoria
            WHERE s.id_trabajador = %s AND s.estado = 'completada'
            ORDER BY s.fecha_finalizacion DESC
            LIMIT 5
        """, (id_persona,))
        historial_trabajos = cursor.fetchall()
        for h in historial_trabajos:
            if h.get('fecha_finalizacion') and hasattr(h['fecha_finalizacion'], 'strftime'):
                h['fecha_finalizacion'] = (h['fecha_finalizacion'] - timedelta(hours=5)).strftime('%d/%m/%Y')
            else:
                h['fecha_finalizacion'] = ''

        ctx = {
            "request":            request,
            "perfil":             perfil,
            "iniciales":          iniciales,
            "verificado":         verificado,
            "servicios":          servicios,
            "calificacion":       calificacion,
            "total_resenas":      total_resenas,
            "trabajos_completados": trabajos_completados,
            "anios_experiencia":  anios_experiencia,
            "horario_texto":      horario_texto,
            "dias_texto":         dias_texto,
            "esta_disponible":    esta_disponible,
            "tarifa_principal":   tarifa_principal,
            "tiene_ayudante":     tiene_ayudante,
            "costo_ayudante":     costo_ayudante,
            "resenas":            resenas,
            "miembro_desde":      miembro_desde,
            "historial_trabajos": historial_trabajos,
        }
        return templates.TemplateResponse("trabajadores/perfil_publico.html", ctx)

    except HTTPException:
        raise
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if conexion and conexion.is_connected():
            conexion.close()

@router.get("/perfil_api/{id_persona}")
def api_perfil_publico(id_persona: int):
    """API del perfil público del trabajador (mantiene compatibilidad)"""
    conexion = conectar_bd()
    try:
        cursor = conexion.cursor(dictionary=True)
        cursor.execute("""
            SELECT p.id_persona, p.nombre_completo, p.ciudad, p.departamento,
                   tp.telefono, ep.anios_experiencia
            FROM personas p
            LEFT JOIN telefono_persona tp ON p.id_persona = tp.id_persona
            LEFT JOIN experiencia_persona ep ON p.id_persona = ep.id_persona
            WHERE p.id_persona = %s
        """, (id_persona,))
        perfil = cursor.fetchone()
        if not perfil:
            return JSONResponse({"error": "No encontrado"}, status_code=404)
        for k, v in perfil.items():
            if v is None: perfil[k] = ''
            elif hasattr(v, 'isoformat'): perfil[k] = str(v)

        cursor.execute("""
            SELECT categoria, descripcion, valor_hora, anios_experiencia
            FROM servicios_persona WHERE id_persona = %s
        """, (id_persona,))
        servicios = cursor.fetchall()
        for s in servicios:
            for k, v in s.items():
                if hasattr(v, '__float__'): s[k] = float(v)
                elif v is None: s[k] = ''
        perfil['servicios'] = servicios

        cursor.execute("""
            SELECT ROUND(AVG(puntuacion),1) as promedio, COUNT(*) as total
            FROM calificaciones WHERE id_trabajador = %s
        """, (id_persona,))
        cal = cursor.fetchone()
        perfil['calificacion'] = float(cal['promedio']) if cal and cal['promedio'] else 0
        perfil['total_calificaciones'] = int(cal['total']) if cal else 0

        cursor.execute("""
            SELECT puntuacion, comentario, fecha_calificacion
            FROM calificaciones WHERE id_trabajador = %s
            ORDER BY fecha_calificacion DESC LIMIT 5
        """, (id_persona,))
        resenas = cursor.fetchall()
        for r in resenas:
            if r.get('fecha_calificacion') and hasattr(r['fecha_calificacion'], 'isoformat'):
                from datetime import timedelta
                r['fecha_calificacion'] = (r['fecha_calificacion'] - timedelta(hours=5)).strftime('%d/%m/%Y')
        perfil['resenas'] = resenas

        return JSONResponse(perfil)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        if conexion and conexion.is_connected():
            conexion.close()

@router.get("/stats/{id_trabajador}")
def obtener_stats_trabajador(id_trabajador: int):
    """Estadísticas del panel del trabajador"""
    conexion = conectar_bd()
    try:
        cursor = conexion.cursor(dictionary=True)

        # Total solicitudes recibidas
        cursor.execute("""
            SELECT COUNT(*) as total FROM solicitudes_servicio
            WHERE id_trabajador = %s
        """, (id_trabajador,))
        total_solicitudes = cursor.fetchone()['total']

        # Solicitudes pendientes
        cursor.execute("""
            SELECT COUNT(*) as total FROM solicitudes_servicio
            WHERE id_trabajador = %s AND estado = 'pendiente'
        """, (id_trabajador,))
        pendientes = cursor.fetchone()['total']

        # Solicitudes completadas (total histórico)
        cursor.execute("""
            SELECT COUNT(*) as total FROM solicitudes_servicio
            WHERE id_trabajador = %s AND estado = 'completada'
        """, (id_trabajador,))
        completadas_total = cursor.fetchone()['total']

        # Solicitudes completadas este mes
        cursor.execute("""
            SELECT COUNT(*) as total FROM solicitudes_servicio
            WHERE id_trabajador = %s AND estado = 'completada'
            AND MONTH(fecha_finalizacion) = MONTH(CURDATE())
            AND YEAR(fecha_finalizacion) = YEAR(CURDATE())
        """, (id_trabajador,))
        completadas_mes = cursor.fetchone()['total']

        # Calificación promedio
        cursor.execute("""
            SELECT AVG(puntuacion) as promedio, COUNT(*) as total
            FROM calificaciones
            WHERE id_trabajador = %s
        """, (id_trabajador,))
        cal = cursor.fetchone()
        calificacion = round(float(cal['promedio']), 1) if cal['promedio'] else 0
        total_cal = cal['total']

        # Estado de disponibilidad
        cursor.execute("""
            SELECT disponible FROM disponibilidad
            WHERE id_persona = %s LIMIT 1
        """, (id_trabajador,))
        disp = cursor.fetchone()
        disponible = bool(disp and disp['disponible']) if disp else False

        # Últimas 3 solicitudes
        cursor.execute("""
            SELECT s.id_solicitud, s.titulo, s.estado, s.fecha_solicitud,
                   c.nombre_completo as nombre_cliente,
                   cat.nombre_categoria
            FROM solicitudes_servicio s
            LEFT JOIN clientes c ON s.id_cliente = c.id_cliente
            LEFT JOIN categorias_servicio cat ON s.id_categoria = cat.id_categoria
            WHERE s.id_trabajador = %s
            ORDER BY s.fecha_solicitud DESC
            LIMIT 3
        """, (id_trabajador,))
        ultimas = cursor.fetchall()
        from datetime import timedelta
        for u in ultimas:
            for k, v in u.items():
                if hasattr(v, 'isoformat'):
                    u[k] = (v - timedelta(hours=5)).strftime('%d/%m %H:%M')
                elif v is None:
                    u[k] = ''

        return JSONResponse({
            "total_solicitudes": total_solicitudes,
            "pendientes": pendientes,
            "completadas": completadas_total,
            "completadas_mes": completadas_mes,
            "calificacion": calificacion,
            "total_calificaciones": total_cal,
            "disponible": disponible,
            "ultimas_solicitudes": ultimas
        })
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        if conexion and conexion.is_connected():
            conexion.close()

@router.get("/registro", response_class=HTMLResponse)
def mostrar_registro_trabajador(request: Request):
    return templates.TemplateResponse("trabajadores/registro_trabajador.html", {"request": request})

@router.post("/crear")
async def crear_trabajador(
    request: Request,
    tipo_documento: int = Form(...),
    numero_documento: str = Form(...),
    genero: int = Form(...),
    nombre_completo: str = Form(...),
    fecha_nacimiento: str = Form(None),
    nacionalidad: str = Form(None),
    ciudad: str = Form(...),
    departamento: str = Form(None),
    codigo_dane: str = Form(...),
    celular: str = Form(...),
    correo: str = Form(None),
    habilidades_tipo: int = Form(...),
    disponibilidad: int = Form(...),
    disponibilidad_dias: int = Form(...),
    foto_identificacion: UploadFile = File(...),
    foto_antecedentes: UploadFile = File(...),
    recomendaciones_archivo: UploadFile = File(...),
    antecedentes: str = Form(None),
    recomendaciones: str = Form(None),
    acepta_terminos: str = Form(None),
    permisos_ubicacion: str = Form(None),
    medio_pago: str = Form(None),
    banco: str = Form(None),
    tipo_cuenta: str = Form(None),
    numero_cuenta: str = Form(None),
    titular_cuenta: str = Form(None),
    medio_pago_principal: str = Form(None),
    arl: str = Form(None),
    eps: str = Form(None)
):
    """Crear nuevo trabajador con todos sus datos"""
    import os
    from datetime import datetime
    from werkzeug.utils import secure_filename
    
    UPLOAD_FOLDER = "static/uploads"
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    
    conexion = None
    cursor = None
    
    try:
        # Validar que el documento no exista
        conexion = conectar_bd()
        cursor = conexion.cursor()
        
        cursor.execute("SELECT COUNT(*) FROM personas WHERE numero_documento = %s", (numero_documento,))
        if cursor.fetchone()[0] > 0:
            return {"error": f"El documento {numero_documento} ya está registrado"}
        
        # Leer archivos en memoria (para guardar en BD)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Foto identificación
        foto_bytes = await foto_identificacion.read()
        foto_filename = f"foto_{timestamp}_{foto_identificacion.filename}"
        foto_tipo = foto_identificacion.content_type or 'image/jpeg'
        # Guardar también en disco si es posible
        try:
            foto_path = os.path.join(UPLOAD_FOLDER, foto_filename)
            with open(foto_path, "wb") as f:
                f.write(foto_bytes)
        except: pass

        # Antecedentes
        antecedentes_bytes = await foto_antecedentes.read()
        antecedentes_filename = f"antecedentes_{timestamp}_{foto_antecedentes.filename}" if foto_antecedentes.filename else None
        antecedentes_tipo = foto_antecedentes.content_type or 'application/pdf'
        try:
            if antecedentes_filename:
                ant_path = os.path.join(UPLOAD_FOLDER, antecedentes_filename)
                with open(ant_path, "wb") as f:
                    f.write(antecedentes_bytes)
        except: pass

        # Recomendaciones
        recomend_bytes = await recomendaciones_archivo.read()
        recomend_filename = f"recom_{timestamp}_{recomendaciones_archivo.filename}" if recomendaciones_archivo.filename else None
        recomend_tipo = recomendaciones_archivo.content_type or 'application/pdf'
        try:
            if recomend_filename:
                rec_path = os.path.join(UPLOAD_FOLDER, recomend_filename)
                with open(rec_path, "wb") as f:
                    f.write(recomend_bytes)
        except: pass
        
        # Iniciar transacción
        conexion.autocommit = False
        
        # 1. Insertar persona
        cursor.execute("""
            INSERT INTO personas 
            (id_tipo_documento, numero_documento, id_genero, nombre_completo, ciudad, codigo_dane, fecha_nacimiento, nacionalidad, registrado_por, departamento)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (tipo_documento, numero_documento, genero, nombre_completo, ciudad.title() if ciudad else '', codigo_dane, fecha_nacimiento, nacionalidad, nombre_completo, departamento.title() if departamento else ''))        
        id_persona = cursor.lastrowid
        
        # 2. Insertar teléfono
        cursor.execute("""
            INSERT INTO telefono_persona (id_persona, telefono)
            VALUES (%s, %s)
        """, (id_persona, celular))
        
        # 3. Insertar correo (opcional)
        if correo:
            cursor.execute("""
                INSERT INTO correo_persona (id_persona, correo)
                VALUES (%s, %s)
            """, (id_persona, correo))
        
        # 4. Obtener servicios del formulario
        form_data = await request.form()
        servicios_categoria = form_data.getlist('servicio_categoria[]')
        servicios_desc = form_data.getlist('habilidad_desc[]')
        servicios_exp = form_data.getlist('habilidad_exp[]')
        servicios_valor = form_data.getlist('habilidad_valor[]')
        tiene_ayudante_list = form_data.getlist('tiene_ayudante[]')
        costo_ayudante_list = form_data.getlist('costo_ayudante[]')
        
        experiencias_validas = []
        todos_servicios = []
        
        for i in range(len(servicios_desc)):
            if servicios_desc[i].strip():
                categoria = servicios_categoria[i] if i < len(servicios_categoria) else 'Otro'
                exp_valor = float(servicios_exp[i]) if servicios_exp[i] else 0
                valor_hora = float(servicios_valor[i]) if servicios_valor[i] else 0
                tiene_ayudante = int(tiene_ayudante_list[i]) if i < len(tiene_ayudante_list) else 0
                costo_ayudante = float(costo_ayudante_list[i]) if i < len(costo_ayudante_list) and costo_ayudante_list[i] and tiene_ayudante == 1 else None
                
                cursor.execute("""
                    INSERT INTO servicios_persona 
                    (id_persona, categoria, descripcion, anios_experiencia, valor_hora, tiene_ayudante, costo_ayudante)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (id_persona, categoria, servicios_desc[i].strip(), exp_valor, valor_hora, tiene_ayudante, costo_ayudante))
                
                experiencias_validas.append(exp_valor)
                todos_servicios.append(f"{categoria}: {servicios_desc[i].strip()}")
        
        # 5. Insertar experiencia general
        max_experiencia = max(experiencias_validas) if experiencias_validas else 0
        resumen_servicios = ' | '.join(todos_servicios)
        
        cursor.execute("""
            INSERT INTO experiencia_persona (id_persona, anios_experiencia, descripcion)
            VALUES (%s, %s, %s)
        """, (id_persona, int(max_experiencia), resumen_servicios))
        
        # 6. Insertar disponibilidad
        cursor.execute("""
            INSERT INTO disponibilidad (id_persona, id_horario, id_dias)
            VALUES (%s, %s, %s)
        """, (id_persona, disponibilidad, disponibilidad_dias))
        
        # 7. Insertar detalles
        acepta_terminos_val = 1 if acepta_terminos == 'on' else 0
        permisos_ubicacion_val = 1 if permisos_ubicacion == 'on' else 0
        
        cursor.execute("""
            INSERT INTO detalles_persona 
            (id_persona, id_servicio_tipo, tareas, antecedentes_pdf, foto_identificacion, 
             acepta_terminos, permisos_ubicacion, recomendaciones, recomendaciones_archivo,
             foto_identificacion_data, foto_identificacion_tipo,
             antecedentes_data, antecedentes_tipo,
             recomendaciones_data, recomendaciones_tipo,
             medio_pago, banco, tipo_cuenta, numero_cuenta, titular_cuenta, medio_pago_principal,
             arl, eps)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (id_persona, habilidades_tipo, resumen_servicios, antecedentes_filename, 
              foto_filename, acepta_terminos_val, permisos_ubicacion_val, 
              recomendaciones or '', recomend_filename,
              foto_bytes, foto_tipo,
              antecedentes_bytes, antecedentes_tipo,
              recomend_bytes, recomend_tipo,
              medio_pago or '', banco or '', tipo_cuenta or '',
              numero_cuenta or '', titular_cuenta or '',
              medio_pago_principal or medio_pago or '',
              arl or '', eps or ''))
        
        conexion.commit()

        return {
            "mensaje": f"✅ Registro exitoso: {nombre_completo} fue registrado",
            "id_persona": id_persona,
            "redirect": f"/trabajador/crear_password?id_persona={id_persona}"
        }
        
    except Exception as e:
        if conexion:
            conexion.rollback()
        import traceback
        error_detalle = traceback.format_exc()
        print(f"ERROR EN REGISTRO: {error_detalle}")
        return {"error": f"Error: {str(e)}"}
    
    finally:
        if cursor:
            cursor.close()
        if conexion and conexion.is_connected():
            conexion.close()


# ============================================
# API ENDPOINTS
# ============================================

@router.get("/solicitudes_pendientes")
def listar_solicitudes_pendientes(id_categoria: int = None):
    conexion = conectar_bd()
    cursor = conexion.cursor(dictionary=True)
    
    if id_categoria:
        sql = """
        SELECT s.*, c.nombre_completo as nombre_cliente, cat.nombre_categoria
        FROM solicitudes_servicio s
        INNER JOIN clientes c ON s.id_cliente = c.id_cliente
        INNER JOIN categorias_servicio cat ON s.id_categoria = cat.id_categoria
        WHERE s.estado = 'pendiente' AND s.id_categoria = %s
        ORDER BY s.fecha_solicitud DESC
        """
        cursor.execute(sql, (id_categoria,))
    else:
        sql = """
        SELECT s.*, c.nombre_completo as nombre_cliente, cat.nombre_categoria
        FROM solicitudes_servicio s
        INNER JOIN clientes c ON s.id_cliente = c.id_cliente
        INNER JOIN categorias_servicio cat ON s.id_categoria = cat.id_categoria
        WHERE s.estado = 'pendiente'
        ORDER BY s.fecha_solicitud DESC
        """
        cursor.execute(sql)
    
    solicitudes = cursor.fetchall()
    cursor.close()
    conexion.close()
    
    return {"solicitudes": solicitudes}

@router.post("/solicitud/aceptar")
def aceptar_solicitud(
    id_solicitud: int = Form(...),
    id_trabajador: int = Form(...)
):
    conexion = conectar_bd()
    cursor = conexion.cursor()
    
    sql = """
    UPDATE solicitudes_servicio 
    SET id_trabajador = %s, estado = 'aceptada', fecha_aceptacion = %s
    WHERE id_solicitud = %s AND estado = 'pendiente'
    """
    
    cursor.execute(sql, (id_trabajador, datetime.now(), id_solicitud))
    conexion.commit()
    
    if cursor.rowcount == 0:
        cursor.close()
        conexion.close()
        return {"error": "La solicitud ya fue aceptada o no existe"}
    
    cursor.close()
    conexion.close()
    
    return {"mensaje": "Solicitud aceptada exitosamente"}

@router.post("/servicio/iniciar")
def iniciar_servicio(id_solicitud: int = Form(...)):
    conexion = conectar_bd()
    cursor = conexion.cursor()
    
    sql = """
    UPDATE solicitudes_servicio 
    SET estado = 'en_proceso', fecha_inicio = %s
    WHERE id_solicitud = %s AND estado = 'aceptada'
    """
    
    cursor.execute(sql, (datetime.now(), id_solicitud))
    conexion.commit()
    cursor.close()
    conexion.close()
    
    return {"mensaje": "Servicio iniciado"}

@router.post("/servicio/completar")
def completar_servicio(
    id_solicitud: int = Form(...),
    precio_final: float = Form(0)
):
    """DESHABILITADO: Solo el cliente puede marcar como completado desde su panel"""
    return JSONResponse(
        {"error": "El servicio solo puede ser marcado como completado por el cliente"},
        status_code=403
    )

@router.post("/calificar-cliente")
def calificar_cliente(
    id_solicitud:  int  = Form(...),
    id_trabajador: int  = Form(...),
    id_cliente:    int  = Form(...),
    puntuacion:    int  = Form(...),
    comentario:    str  = Form(None),
    tags:          str  = Form(None)
):
    """El trabajador califica al cliente tras completar el servicio."""
    if not (1 <= puntuacion <= 5):
        return JSONResponse({"error": "La puntuación debe estar entre 1 y 5"}, status_code=400)

    conexion = conectar_bd()
    cursor   = conexion.cursor(dictionary=True)
    try:
        # Buscar la solicitud completada — validar solo por id_solicitud
        # y verificar que id_trabajador o id_cliente coincidan con lo enviado
        cursor.execute("""
            SELECT id_solicitud, id_trabajador, id_cliente FROM solicitudes_servicio
            WHERE id_solicitud = %s
              AND estado = 'completada'
            LIMIT 1
        """, (id_solicitud,))
        sol = cursor.fetchone()

        if not sol:
            return JSONResponse(
                {"error": "Solo puedes calificar un servicio que hayas completado con este cliente"},
                status_code=403
            )

        # Usar los IDs reales de la BD (no confiar en el frontend)
        real_trabajador = sol['id_trabajador']
        real_cliente    = sol['id_cliente']

        # Evitar duplicado
        cursor.execute("""
            SELECT id_calificacion FROM calificaciones
            WHERE id_solicitud    = %s
              AND id_trabajador   = %s
              AND tipo_calificacion = 'trabajador_a_cliente'
        """, (id_solicitud, real_trabajador))
        if cursor.fetchone():
            return JSONResponse({"mensaje": "Ya calificaste a este cliente"})

        cursor.execute("""
            INSERT INTO calificaciones
                (id_solicitud, id_cliente, id_trabajador, tipo_calificacion, puntuacion, comentario, tags)
            VALUES (%s, %s, %s, 'trabajador_a_cliente', %s, %s, %s)
        """, (id_solicitud, real_cliente, real_trabajador, puntuacion, comentario, tags))
        conexion.commit()
        return JSONResponse({"mensaje": "¡Gracias! Calificación enviada."})
    except Exception as e:
        conexion.rollback()
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        cursor.close()
        conexion.close()


@router.get("/historial/{id_trabajador}")
def historial_trabajador(id_trabajador: int):
    conexion = conectar_bd()
    cursor = conexion.cursor(dictionary=True)
    
    cursor.execute("""
        SELECT s.*, c.nombre_completo as nombre_cliente, cat.nombre_categoria
        FROM solicitudes_servicio s
        INNER JOIN clientes c ON s.id_cliente = c.id_cliente
        INNER JOIN categorias_servicio cat ON s.id_categoria = cat.id_categoria
        WHERE s.id_trabajador = %s
        ORDER BY s.fecha_solicitud DESC
    """, (id_trabajador,))
    
    historial = cursor.fetchall()
    cursor.close()
    conexion.close()
    
    return {"historial": historial}

@router.get("/disponibles/{id_categoria}")
def trabajadores_disponibles(id_categoria: int):
    """Trabajadores disponibles por categoría — usa tablas reales del proyecto"""
    conexion = conectar_bd()
    cursor = conexion.cursor(dictionary=True)
    try:
        cursor.execute("""
            SELECT 
                p.id_persona,
                p.nombre_completo,
                p.ciudad,
                sp.anios_experiencia,
                sp.categoria,
                COALESCE(AVG(cal.puntuacion), 0) as calificacion,
                COUNT(DISTINCT sol.id_solicitud) as servicios_completados
            FROM personas p
            INNER JOIN servicios_persona sp ON p.id_persona = sp.id_persona
            LEFT JOIN calificaciones cal ON p.id_persona = cal.id_trabajador
            LEFT JOIN solicitudes_servicio sol ON p.id_persona = sol.id_trabajador AND sol.estado = 'completada'
            LEFT JOIN disponibilidad d ON p.id_persona = d.id_persona
            INNER JOIN categorias_servicio cat ON sp.categoria = cat.nombre_categoria
            WHERE cat.id_categoria = %s
            AND (p.estado = 'activo' OR p.estado IS NULL)
            AND (d.disponible = 1 OR d.disponible IS NULL)
            GROUP BY p.id_persona, p.nombre_completo, p.ciudad, sp.anios_experiencia, sp.categoria
            ORDER BY calificacion DESC
        """, (id_categoria,))
        trabajadores = cursor.fetchall()
        for t in trabajadores:
            for k, v in t.items():
                if hasattr(v, '__float__'): t[k] = float(v)
                elif v is None: t[k] = ''
        return JSONResponse({"trabajadores": trabajadores})
    except Exception as e:
        return JSONResponse({"error": str(e), "trabajadores": []}, status_code=500)
    finally:
        cursor.close()
        conexion.close()


# ============================================
# AUTENTICACIÓN Y ADMINISTRACIÓN
# ============================================

# Contraseña de administrador - se lee desde variable de entorno en Railway
import os
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "admin123")
ADMIN_TOKEN_SECRET = os.getenv("ADMIN_TOKEN", "talenthub_admin_2026_secret")

def verificar_admin(request: Request) -> bool:
    """Verifica si la request tiene sesión admin válida"""
    token = request.cookies.get("admin_session")
    return token == ADMIN_TOKEN_SECRET

def requiere_admin(request: Request):
    """Redirige a login si no es admin"""
    if not verificar_admin(request):
        return RedirectResponse(url="/trabajador/admin/login", status_code=302)
    return None

@router.get("/admin/login", response_class=HTMLResponse)
def mostrar_admin_login(request: Request):
    return templates.TemplateResponse("trabajadores/admin_login.html", {"request": request})

@router.post("/admin/login")
async def admin_login(response: Response, password: str = Form(...)):
    """Login de administrador — establece cookie de sesión"""
    try:
        if password == ADMIN_PASSWORD:
            resp = JSONResponse({
                "success": True,
                "mensaje": "Inicio de sesión exitoso"
            })
            resp.set_cookie(
                key="admin_session",
                value=ADMIN_TOKEN_SECRET,
                httponly=True,
                max_age=86400 * 7,  # 7 días
                samesite="lax"
            )
            return resp
        else:
            return JSONResponse({
                "success": False,
                "error": "Contraseña incorrecta"
            }, status_code=401)
    except Exception as e:
        return JSONResponse({
            "success": False,
            "error": f"Error: {str(e)}"
        }, status_code=500)

@router.get("/admin/logout")
def admin_logout():
    """Cerrar sesión de administrador"""
    resp = RedirectResponse(url="/", status_code=302)
    resp.delete_cookie("admin_session")
    return resp

@router.get("/registros", response_class=HTMLResponse)
def ver_registros(request: Request):
    """Página para ver todos los registros (solo admin)"""
    redir = requiere_admin(request)
    if redir: return redir
    return templates.TemplateResponse("trabajadores/registros.html", {"request": request})

@router.get("/registros/test")
def test_registros():
    """Endpoint de prueba simple"""
    conexion = conectar_bd()
    if not conexion:
        return {"error": "No se pudo conectar a la base de datos"}
    
    try:
        cursor = conexion.cursor(dictionary=True)
        cursor.execute("SELECT COUNT(*) as total FROM personas")
        total = cursor.fetchone()
        
        cursor.execute("SELECT * FROM personas LIMIT 5")
        registros = cursor.fetchall()
        
        # Convertir fechas a string
        for reg in registros:
            if reg.get('fecha_registro'):
                reg['fecha_registro'] = str(reg['fecha_registro'])
        
        return {
            "total": total['total'],
            "registros": registros,
            "mensaje": "Conexión exitosa"
        }
    except Exception as e:
        return {"error": str(e)}
    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()

@router.get("/registros/listar")
def listar_registros(request: Request):
    """API para obtener todos los registros de trabajadores"""
    if not verificar_admin(request):
        return JSONResponse({"error": "No autorizado", "registros": []}, status_code=401)
    conexion = conectar_bd()
    if not conexion:
        return JSONResponse({"error": "Error de conexión", "registros": []}, status_code=500)
    
    try:
        cursor = conexion.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT 
                p.id_persona,
                p.numero_documento,
                p.nombre_completo,
                p.ciudad,
                p.departamento,
                p.codigo_dane,
                p.id_genero,
                p.fecha_nacimiento,
                p.nacionalidad,
                p.fecha_registro,
                p.estado
            FROM personas p
            WHERE p.estado = 'activo' OR p.estado IS NULL
            ORDER BY p.id_persona DESC
        """)
        
        registros = cursor.fetchall()
        
        generos = {1: 'Masculino', 2: 'Femenino', 3: 'Otro', 4: 'Masculino', 5: 'Femenino', 6: 'Otro'}
        horarios = {7: '8am-12pm', 8: '2pm-6pm', 9: '6pm-10pm', 10: '24 horas'}
        dias = {11: 'Entre Semana', 12: 'Toda la Semana'}

        for reg in registros:
            for key, value in reg.items():
                if value is None:
                    reg[key] = ''
                elif hasattr(value, 'isoformat'):
                    # Convertir UTC a hora Colombia (UTC-5)
                    from datetime import timedelta
                    valor_col = value - timedelta(hours=5)
                    reg[key] = valor_col.strftime('%Y-%m-%d %H:%M')
                elif isinstance(value, (int, float)):
                    reg[key] = float(value) if isinstance(value, float) else int(value)
                else:
                    reg[key] = str(value)

            # Género desde id_genero
            id_gen = reg.get('id_genero', 0)
            reg['genero'] = generos.get(int(id_gen) if id_gen else 0, 'N/A')
            reg['tipo_documento'] = 'CC'
            reg['telefono'] = 'N/A'
            reg['correo'] = ''
            reg['anios_experiencia'] = 0
            reg['servicio_tipo'] = 'N/A'
            reg['foto_identificacion'] = ''
            reg['antecedentes_pdf'] = ''
            reg['recomendaciones'] = ''
            reg['recomendaciones_archivo'] = ''
            reg['arl'] = ''
            reg['eps'] = ''
            reg['horario'] = 'N/A'
            reg['dias_disponibles'] = 'N/A'
            reg['servicios'] = []
            reg['registrado_por'] = 'Sistema'

            try:
                cursor.execute("SELECT telefono FROM telefono_persona WHERE id_persona = %s LIMIT 1", (reg['id_persona'],))
                tel = cursor.fetchone()
                if tel: reg['telefono'] = str(tel['telefono'])

                cursor.execute("SELECT correo FROM correo_persona WHERE id_persona = %s LIMIT 1", (reg['id_persona'],))
                email = cursor.fetchone()
                if email: reg['correo'] = str(email['correo'])

                cursor.execute("""
                    SELECT foto_identificacion, antecedentes_pdf, recomendaciones, recomendaciones_archivo,
                           arl, eps, medio_pago, medio_pago_principal, banco, numero_cuenta,
                           (foto_identificacion_data IS NOT NULL AND LENGTH(foto_identificacion_data) > 0) as tiene_foto,
                           (antecedentes_data IS NOT NULL AND LENGTH(antecedentes_data) > 0) as tiene_ant,
                           (recomendaciones_data IS NOT NULL AND LENGTH(recomendaciones_data) > 0) as tiene_rec
                    FROM detalles_persona WHERE id_persona = %s LIMIT 1
                """, (reg['id_persona'],))
                detalles = cursor.fetchone()
                if detalles:
                    reg['foto_identificacion'] = str(detalles.get('foto_identificacion') or '')
                    reg['antecedentes_pdf'] = str(detalles.get('antecedentes_pdf') or '')
                    reg['recomendaciones'] = str(detalles.get('recomendaciones') or '')
                    reg['recomendaciones_archivo'] = str(detalles.get('recomendaciones_archivo') or '')
                    reg['medio_pago'] = str(detalles.get('medio_pago') or '')
                    reg['medio_pago_principal'] = str(detalles.get('medio_pago_principal') or '')
                    reg['arl'] = str(detalles.get('arl') or '')
                    reg['eps'] = str(detalles.get('eps') or '')
                    # URLs para servir desde BD
                    if detalles.get('tiene_foto'):
                        reg['foto_identificacion_url'] = f"/trabajador/archivo/{reg['id_persona']}/foto"
                    if detalles.get('tiene_ant'):
                        reg['antecedentes_url'] = f"/trabajador/archivo/{reg['id_persona']}/antecedentes"
                    if detalles.get('tiene_rec'):
                        reg['recomendaciones_url'] = f"/trabajador/archivo/{reg['id_persona']}/recomendaciones"

                # Disponibilidad
                cursor.execute("SELECT id_horario, id_dias FROM disponibilidad WHERE id_persona = %s LIMIT 1", (reg['id_persona'],))
                disp = cursor.fetchone()
                if disp:
                    reg['horario'] = horarios.get(disp.get('id_horario'), 'N/A')
                    reg['dias_disponibles'] = dias.get(disp.get('id_dias'), 'N/A')

                cursor.execute("""
                    SELECT categoria, descripcion, anios_experiencia, valor_hora, tiene_ayudante, costo_ayudante
                    FROM servicios_persona WHERE id_persona = %s
                """, (reg['id_persona'],))
                servicios = cursor.fetchall()

                max_exp = 0
                servicios_proc = []
                for serv in servicios:
                    sp = {}
                    for k, v in serv.items():
                        if hasattr(v, '__float__'): sp[k] = float(v)
                        elif v is None: sp[k] = 0 if k in ['valor_hora','anios_experiencia','tiene_ayudante','costo_ayudante'] else ''
                        else: sp[k] = v
                    servicios_proc.append(sp)
                    if sp.get('anios_experiencia', 0) > max_exp:
                        max_exp = sp['anios_experiencia']

                reg['servicios'] = servicios_proc
                reg['anios_experiencia'] = int(max_exp)
                if servicios_proc:
                    reg['servicio_tipo'] = servicios_proc[0]['categoria']

            except Exception as ex:
                print(f"Error datos adicionales persona {reg['id_persona']}: {ex}")
                pass
        
        return JSONResponse({"registros": registros})
        
    except Exception as e:
        import traceback
        print(f"Error en /registros/listar: {traceback.format_exc()}")
        return JSONResponse({"error": str(e), "registros": []}, status_code=500)
    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()

@router.get("/dashboard", response_class=HTMLResponse)
def mostrar_dashboard(request: Request):
    """Dashboard principal del administrador"""
    redir = requiere_admin(request)
    if redir: return redir
    return templates.TemplateResponse("trabajadores/dashboard.html", {"request": request})

@router.get("/estadisticas")
def obtener_estadisticas(request: Request):
    """API para obtener estadísticas completas del dashboard (solo admin)"""
    if not verificar_admin(request):
        return JSONResponse({"error": "No autorizado"}, status_code=401)
    conexion = conectar_bd()
    if not conexion:
        return JSONResponse({"error": "Error de conexión"}, status_code=500)
    
    try:
        cursor = conexion.cursor(dictionary=True)
        
        # Totales generales
        cursor.execute("SELECT COUNT(*) as total FROM personas WHERE estado = 'activo' OR estado IS NULL")
        total_activos = cursor.fetchone()['total']
        
        cursor.execute("SELECT COUNT(*) as total FROM personas WHERE estado = 'eliminado'")
        total_eliminados = cursor.fetchone()['total']

        cursor.execute("SELECT COUNT(*) as total FROM clientes WHERE estado = 'activo'")
        total_clientes = cursor.fetchone()['total']
        
        cursor.execute("SELECT COUNT(*) as total FROM personas")
        total_todos = cursor.fetchone()['total']

        # Registros hoy
        cursor.execute("SELECT COUNT(*) as total FROM personas WHERE DATE(fecha_registro) = CURDATE()")
        row = cursor.fetchone()
        registros_hoy = row['total'] if row else 0

        # Registros esta semana
        cursor.execute("SELECT COUNT(*) as total FROM personas WHERE fecha_registro >= DATE_SUB(NOW(), INTERVAL 7 DAY)")
        row = cursor.fetchone()
        registros_semana = row['total'] if row else 0

        # Por categoría de servicio
        cursor.execute("""
            SELECT categoria, COUNT(*) as total 
            FROM servicios_persona 
            GROUP BY categoria 
            ORDER BY total DESC
            LIMIT 10
        """)
        por_categoria = cursor.fetchall()

        # Por ciudad (top 10)
        cursor.execute("""
            SELECT ciudad, COUNT(*) as total 
            FROM personas 
            WHERE (estado = 'activo' OR estado IS NULL) AND ciudad IS NOT NULL AND ciudad != ''
            GROUP BY ciudad 
            ORDER BY total DESC
            LIMIT 10
        """)
        por_ciudad = cursor.fetchall()

        # Por departamento (top 10)
        cursor.execute("""
            SELECT departamento, COUNT(*) as total 
            FROM personas 
            WHERE (estado = 'activo' OR estado IS NULL) AND departamento IS NOT NULL AND departamento != ''
            GROUP BY departamento 
            ORDER BY total DESC
            LIMIT 10
        """)
        por_departamento = cursor.fetchall()

        # Registros por mes (últimos 6 meses)
        cursor.execute("""
            SELECT DATE_FORMAT(fecha_registro, '%Y-%m') as mes, COUNT(*) as total
            FROM personas
            WHERE fecha_registro >= DATE_SUB(NOW(), INTERVAL 6 MONTH)
            GROUP BY mes ORDER BY mes ASC
        """)
        por_mes = cursor.fetchall()

        # Tarifa promedio
        cursor.execute("SELECT AVG(valor_hora) as promedio FROM servicios_persona WHERE valor_hora > 0")
        row = cursor.fetchone()
        tarifa_promedio = float(row['promedio']) if row and row['promedio'] else 0

        # Últimos 5 registros
        cursor.execute("""
            SELECT p.id_persona, p.nombre_completo, p.ciudad, p.departamento,
                   p.fecha_registro,
                   GROUP_CONCAT(DISTINCT sp.categoria SEPARATOR ', ') as servicios
            FROM personas p
            LEFT JOIN servicios_persona sp ON p.id_persona = sp.id_persona
            WHERE p.estado = 'activo' OR p.estado IS NULL
            GROUP BY p.id_persona
            ORDER BY p.id_persona DESC
            LIMIT 5
        """)
        ultimos_registros = cursor.fetchall()
        for r in ultimos_registros:
            if r.get('fecha_registro'):
                r['fecha_registro'] = str(r['fecha_registro'])

        # ── MÉTRICAS DE SOLICITUDES ────────────────────────────────

        # Totales por estado
        cursor.execute("""
            SELECT estado, COUNT(*) as total
            FROM solicitudes_servicio
            GROUP BY estado
        """)
        rows = cursor.fetchall()
        sol_por_estado = {r['estado']: r['total'] for r in rows}
        total_solicitudes   = sum(sol_por_estado.values())
        sol_pendientes      = sol_por_estado.get('pendiente',  0)
        sol_aceptadas       = sol_por_estado.get('aceptada',   0)
        sol_en_proceso      = sol_por_estado.get('en_proceso', 0)
        sol_completadas     = sol_por_estado.get('completada', 0)
        sol_canceladas      = sol_por_estado.get('cancelada',  0)

        # Tasa de conversión: completadas / (completadas + canceladas) * 100
        base_conv = sol_completadas + sol_canceladas
        tasa_conversion = round(sol_completadas / base_conv * 100, 1) if base_conv > 0 else 0

        # Solicitudes por día (últimos 14 días)
        cursor.execute("""
            SELECT DATE(fecha_solicitud) as dia, COUNT(*) as total
            FROM solicitudes_servicio
            WHERE fecha_solicitud >= DATE_SUB(CURDATE(), INTERVAL 14 DAY)
            GROUP BY dia ORDER BY dia ASC
        """)
        sol_por_dia = [{"dia": str(r['dia']), "total": r['total']} for r in cursor.fetchall()]

        # Top categorías más solicitadas
        cursor.execute("""
            SELECT COALESCE(cat.nombre_categoria, s.titulo) as categoria, COUNT(*) as total
            FROM solicitudes_servicio s
            LEFT JOIN categorias_servicio cat ON s.id_categoria = cat.id_categoria
            GROUP BY categoria ORDER BY total DESC LIMIT 8
        """)
        top_categorias_sol = cursor.fetchall()

        # Tiempo promedio de respuesta (minutos entre fecha_solicitud y fecha_aceptacion)
        cursor.execute("""
            SELECT ROUND(AVG(TIMESTAMPDIFF(MINUTE, fecha_solicitud, fecha_aceptacion)), 0) as promedio_min
            FROM solicitudes_servicio
            WHERE fecha_aceptacion IS NOT NULL AND fecha_solicitud IS NOT NULL
        """)
        row = cursor.fetchone()
        tiempo_respuesta_min = int(row['promedio_min']) if row and row['promedio_min'] else None

        # Top trabajadores (por trabajos completados)
        cursor.execute("""
            SELECT p.nombre_completo, COUNT(*) as completados,
                   ROUND(AVG(cal.puntuacion), 1) as calificacion
            FROM solicitudes_servicio s
            INNER JOIN personas p ON s.id_trabajador = p.id_persona
            LEFT JOIN calificaciones cal ON s.id_solicitud = cal.id_solicitud
            WHERE s.estado = 'completada'
            GROUP BY s.id_trabajador, p.nombre_completo
            ORDER BY completados DESC LIMIT 5
        """)
        top_trabajadores = cursor.fetchall()
        for t in top_trabajadores:
            if t['calificacion'] is None:
                t['calificacion'] = 0
            else:
                t['calificacion'] = float(t['calificacion'])

        # Solicitudes este mes vs mes anterior
        cursor.execute("""
            SELECT COUNT(*) as total FROM solicitudes_servicio
            WHERE MONTH(fecha_solicitud) = MONTH(CURDATE())
              AND YEAR(fecha_solicitud)  = YEAR(CURDATE())
        """)
        sol_este_mes = cursor.fetchone()['total']

        cursor.execute("""
            SELECT COUNT(*) as total FROM solicitudes_servicio
            WHERE MONTH(fecha_solicitud) = MONTH(DATE_SUB(CURDATE(), INTERVAL 1 MONTH))
              AND YEAR(fecha_solicitud)  = YEAR(DATE_SUB(CURDATE(), INTERVAL 1 MONTH))
        """)
        sol_mes_anterior = cursor.fetchone()['total']

        # Clientes recurrentes (más de 1 solicitud)
        cursor.execute("""
            SELECT COUNT(*) as total FROM (
                SELECT id_cliente FROM solicitudes_servicio
                GROUP BY id_cliente HAVING COUNT(*) > 1
            ) sub
        """)
        clientes_recurrentes = cursor.fetchone()['total']

        return JSONResponse({
            "total_activos":       total_activos,
            "total_profesionales": total_activos,
            "total_eliminados":    total_eliminados,
            "total_clientes":      total_clientes,
            "total_todos":         total_todos,
            "total_categorias":    len(por_categoria),
            "total_ciudades":      len(por_ciudad),
            "registros_hoy":       registros_hoy,
            "registros_semana":    registros_semana,
            "tarifa_promedio":     tarifa_promedio,
            "por_categoria":       por_categoria,
            "por_ciudad":          por_ciudad,
            "por_departamento":    por_departamento,
            "por_mes":             por_mes,
            "ultimos_registros":   ultimos_registros,
            # Solicitudes
            "total_solicitudes":      total_solicitudes,
            "sol_pendientes":         sol_pendientes,
            "sol_aceptadas":          sol_aceptadas,
            "sol_en_proceso":         sol_en_proceso,
            "sol_completadas":        sol_completadas,
            "sol_canceladas":         sol_canceladas,
            "tasa_conversion":        tasa_conversion,
            "sol_por_dia":            sol_por_dia,
            "top_categorias_sol":     top_categorias_sol,
            "tiempo_respuesta_min":   tiempo_respuesta_min,
            "top_trabajadores":       top_trabajadores,
            "sol_este_mes":           sol_este_mes,
            "sol_mes_anterior":       sol_mes_anterior,
            "clientes_recurrentes":   clientes_recurrentes,
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        if conexion and conexion.is_connected():
            conexion.close()

@router.post("/eliminar")
async def eliminar_registro(id_persona: int = Form(...)):
    """Eliminar registro (soft delete)"""
    conexion = conectar_bd()
    if not conexion:
        return JSONResponse({"error": "Error de conexión"}, status_code=500)
    
    try:
        cursor = conexion.cursor()
        
        cursor.execute("""
            UPDATE personas 
            SET estado = 'eliminado', fecha_eliminacion = %s
            WHERE id_persona = %s
        """, (datetime.now(), id_persona))
        
        conexion.commit()
        
        return JSONResponse({"mensaje": "✅ Registro eliminado exitosamente"})
        
    except Exception as e:
        conexion.rollback()
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        if conexion and conexion.is_connected():
            conexion.close()

# ── MAPA ADMIN — ver todos los trabajadores ──────────────────────

@router.get("/admin/mapa", response_class=HTMLResponse)
def mapa_admin(request: Request):
    """Mapa con todos los trabajadores disponibles (admin)"""
    return templates.TemplateResponse("trabajadores/mapa_admin.html", {"request": request})

@router.get("/admin/corregir-ciudades")
def corregir_ciudades(request: Request):
    """Normaliza ciudades y departamentos a Title Case (solo admin)"""
    if not verificar_admin(request):
        return JSONResponse({"error": "No autorizado"}, status_code=401)
    conexion = conectar_bd()
    try:
        cursor = conexion.cursor()
        cursor.execute("""
            UPDATE personas
            SET ciudad = CONCAT(UPPER(SUBSTRING(LOWER(ciudad),1,1)), SUBSTRING(LOWER(ciudad),2)),
                departamento = CONCAT(UPPER(SUBSTRING(LOWER(departamento),1,1)), SUBSTRING(LOWER(departamento),2))
            WHERE ciudad IS NOT NULL AND ciudad != ''
        """)
        afectados = cursor.rowcount
        conexion.commit()
        return JSONResponse({"mensaje": f"Corregidos {afectados} registros", "ok": True})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        if conexion and conexion.is_connected():
            conexion.close()

@router.post("/admin/editar-registro")
def admin_editar_registro(
    request: Request,
    id_persona:   int = Form(...),
    nombre:       str = Form(None),
    ciudad:       str = Form(None),
    departamento: str = Form(None),
    telefono:     str = Form(None),
    correo:       str = Form(None)
):
    """Admin edita datos de un trabajador"""
    if not verificar_admin(request):
        return JSONResponse({"error": "No autorizado"}, status_code=401)
    conexion = conectar_bd()
    try:
        cursor = conexion.cursor()
        if nombre:
            cursor.execute("UPDATE personas SET nombre_completo = %s WHERE id_persona = %s", (nombre, id_persona))
        if ciudad:
            cursor.execute("UPDATE personas SET ciudad = %s WHERE id_persona = %s", (ciudad, id_persona))
        if departamento:
            cursor.execute("UPDATE personas SET departamento = %s WHERE id_persona = %s", (departamento, id_persona))
        if telefono:
            cursor.execute("SELECT 1 FROM telefono_persona WHERE id_persona = %s LIMIT 1", (id_persona,))
            if cursor.fetchone():
                cursor.execute("UPDATE telefono_persona SET telefono = %s WHERE id_persona = %s", (telefono, id_persona))
            else:
                cursor.execute("INSERT INTO telefono_persona (id_persona, telefono) VALUES (%s, %s)", (id_persona, telefono))
        if correo:
            cursor.execute("SELECT 1 FROM correo_persona WHERE id_persona = %s LIMIT 1", (id_persona,))
            if cursor.fetchone():
                cursor.execute("UPDATE correo_persona SET correo = %s WHERE id_persona = %s", (correo, id_persona))
            else:
                cursor.execute("INSERT INTO correo_persona (id_persona, correo) VALUES (%s, %s)", (id_persona, correo))
        conexion.commit()
        return JSONResponse({"mensaje": "Registro actualizado correctamente"})
    except Exception as e:
        conexion.rollback()
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        if conexion and conexion.is_connected():
            conexion.close()

@router.get("/admin/mapa-data")
def mapa_admin_data(request: Request):
    """API: todos los trabajadores con ubicación activa (solo admin)"""
    if not verificar_admin(request):
        return JSONResponse({"trabajadores": [], "error": "No autorizado"}, status_code=401)
    conexion = conectar_bd()
    if not conexion:
        return JSONResponse({"trabajadores": []})
    try:
        cursor = conexion.cursor(dictionary=True)
        cursor.execute("""
            SELECT p.id_persona, p.nombre_completo, p.ciudad,
                   d.latitud, d.longitud, d.disponible, d.ultima_actualizacion,
                   GROUP_CONCAT(DISTINCT sp.categoria SEPARATOR ', ') as categorias
            FROM personas p
            INNER JOIN disponibilidad d ON p.id_persona = d.id_persona
            LEFT JOIN servicios_persona sp ON p.id_persona = sp.id_persona
            WHERE d.disponible = 1
              AND d.latitud IS NOT NULL
              AND d.longitud IS NOT NULL
              AND (p.estado = 'activo' OR p.estado IS NULL)
            GROUP BY p.id_persona, p.nombre_completo, p.ciudad,
                     d.latitud, d.longitud, d.disponible, d.ultima_actualizacion
        """)
        trabajadores = cursor.fetchall()
        from datetime import timedelta
        for t in trabajadores:
            t['nombre'] = t.pop('nombre_completo', '')
            if t.get('ultima_actualizacion') and hasattr(t['ultima_actualizacion'], 'strftime'):
                t['ultima_actualizacion'] = (t['ultima_actualizacion'] - timedelta(hours=5)).strftime('%d/%m %H:%M')
            elif t.get('ultima_actualizacion'):
                t['ultima_actualizacion'] = str(t['ultima_actualizacion'])
            else:
                t['ultima_actualizacion'] = ''
            if t.get('latitud'): t['latitud'] = float(t['latitud'])
            if t.get('longitud'): t['longitud'] = float(t['longitud'])
        return JSONResponse({"trabajadores": trabajadores})
    except Exception as e:
        return JSONResponse({"trabajadores": [], "error": str(e)})
    finally:
        if conexion and conexion.is_connected():
            conexion.close()

@router.get("/exportar-excel")
def exportar_excel(request: Request):
    """Exportar todos los trabajadores activos a Excel (solo admin)"""
    if not verificar_admin(request):
        return JSONResponse({"error": "No autorizado"}, status_code=401)
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    conexion = conectar_bd()
    if not conexion:
        return JSONResponse({"error": "Error de conexión"}, status_code=500)

    try:
        cursor = conexion.cursor(dictionary=True)

        cursor.execute("""
            SELECT
                p.id_persona,
                p.numero_documento,
                p.nombre_completo,
                p.ciudad,
                p.departamento,
                p.fecha_nacimiento,
                p.nacionalidad,
                p.fecha_registro
            FROM personas p
            WHERE p.estado = 'activo' OR p.estado IS NULL
            ORDER BY p.id_persona DESC
        """)
        registros = cursor.fetchall()

        horarios = {7: '8am-12pm', 8: '2pm-6pm', 9: '6pm-10pm', 10: '24 horas'}
        dias_map = {11: 'Entre Semana', 12: 'Toda la Semana'}

        filas = []
        for reg in registros:
            id_p = reg['id_persona']

            # Teléfono
            cursor.execute("SELECT telefono FROM telefono_persona WHERE id_persona = %s LIMIT 1", (id_p,))
            tel = cursor.fetchone()
            telefono = str(tel['telefono']) if tel else ''

            # Correo
            cursor.execute("SELECT correo FROM correo_persona WHERE id_persona = %s LIMIT 1", (id_p,))
            email = cursor.fetchone()
            correo = str(email['correo']) if email else ''

            # Servicios
            cursor.execute("""
                SELECT categoria, descripcion, anios_experiencia, valor_hora
                FROM servicios_persona WHERE id_persona = %s
            """, (id_p,))
            servicios = cursor.fetchall()
            categorias = ', '.join(s['categoria'] for s in servicios if s.get('categoria'))
            valor_hora = max((float(s['valor_hora']) for s in servicios if s.get('valor_hora')), default=0)
            max_exp = max((float(s['anios_experiencia']) for s in servicios if s.get('anios_experiencia')), default=0)

            # Disponibilidad
            cursor.execute("SELECT id_horario, id_dias FROM disponibilidad WHERE id_persona = %s LIMIT 1", (id_p,))
            disp = cursor.fetchone()
            horario = horarios.get(disp['id_horario'], '') if disp else ''
            dias_disp = dias_map.get(disp['id_dias'], '') if disp else ''

            fecha_reg = reg.get('fecha_registro')
            if fecha_reg and hasattr(fecha_reg, 'strftime'):
                from datetime import timedelta
                fecha_reg = (fecha_reg - timedelta(hours=5)).strftime('%d/%m/%Y %H:%M')
            else:
                fecha_reg = str(fecha_reg) if fecha_reg else ''

            fecha_nac = reg.get('fecha_nacimiento')
            fecha_nac = str(fecha_nac) if fecha_nac else ''

            filas.append([
                reg['id_persona'],
                reg.get('numero_documento', ''),
                reg.get('nombre_completo', ''),
                telefono,
                correo,
                reg.get('ciudad', ''),
                reg.get('departamento', ''),
                reg.get('nacionalidad', ''),
                fecha_nac,
                categorias,
                max_exp,
                valor_hora,
                horario,
                dias_disp,
                fecha_reg,
            ])

        cursor.close()
        conexion.close()

        # Crear Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trabajadores"

        encabezados = [
            "ID", "Documento", "Nombre Completo", "Teléfono", "Correo",
            "Ciudad", "Departamento", "Nacionalidad", "Fecha Nacimiento",
            "Categorías", "Años Experiencia", "Valor Hora ($)",
            "Horario", "Días Disponibles", "Fecha Registro"
        ]

        # Estilo encabezado
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.append(encabezados)
        for col_idx, _ in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border

        ws.row_dimensions[1].height = 30

        # Filas de datos
        alt_fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
        data_align = Alignment(vertical="center", wrap_text=False)

        for row_idx, fila in enumerate(filas, start=2):
            ws.append(fila)
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx in range(1, len(encabezados) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.alignment = data_align
                if fill:
                    cell.fill = fill

        # Anchos de columna
        anchos = [6, 14, 28, 14, 26, 16, 16, 14, 14, 30, 10, 14, 12, 16, 18]
        for i, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        from datetime import datetime as dt
        nombre_archivo = f"trabajadores_{dt.now().strftime('%Y%m%d_%H%M')}.xlsx"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"}
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        if conexion and conexion.is_connected():
            conexion.close()
        return JSONResponse({"error": str(e)}, status_code=500)


@router.get("/exportar-metricas")
def exportar_metricas(request: Request):
    """Exportar métricas del dashboard a Excel (solo admin)"""
    if not verificar_admin(request):
        return JSONResponse({"error": "No autorizado"}, status_code=401)
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    from datetime import datetime as dt, timedelta

    conexion = conectar_bd()
    if not conexion:
        return JSONResponse({"error": "Error de conexión"}, status_code=500)

    try:
        cursor = conexion.cursor(dictionary=True)

        # ── Métricas generales ──
        cursor.execute("SELECT COUNT(*) as t FROM personas WHERE estado = 'activo' OR estado IS NULL")
        total_profesionales = cursor.fetchone()['t']
        cursor.execute("SELECT COUNT(*) as t FROM clientes WHERE estado = 'activo' OR estado IS NULL")
        total_clientes = cursor.fetchone()['t']
        cursor.execute("SELECT COUNT(*) as t FROM solicitudes_servicio")
        total_solicitudes = cursor.fetchone()['t']
        cursor.execute("SELECT COUNT(*) as t FROM solicitudes_servicio WHERE estado = 'completada'")
        completadas = cursor.fetchone()['t']
        cursor.execute("SELECT COUNT(*) as t FROM solicitudes_servicio WHERE estado = 'cancelada'")
        canceladas = cursor.fetchone()['t']
        cursor.execute("SELECT COUNT(*) as t FROM solicitudes_servicio WHERE estado = 'pendiente'")
        pendientes = cursor.fetchone()['t']
        cursor.execute("SELECT AVG(valor_hora) as p FROM servicios_persona WHERE valor_hora > 0")
        row = cursor.fetchone()
        tarifa_prom = round(float(row['p']), 0) if row and row['p'] else 0

        # ── Top categorías ──
        cursor.execute("""
            SELECT COALESCE(cat.nombre_categoria, 'Otro') as categoria, COUNT(*) as total
            FROM solicitudes_servicio s
            LEFT JOIN categorias_servicio cat ON s.id_categoria = cat.id_categoria
            GROUP BY categoria ORDER BY total DESC LIMIT 15
        """)
        top_categorias = cursor.fetchall()

        # ── Top trabajadores ──
        cursor.execute("""
            SELECT p.nombre_completo, COUNT(*) as trabajos,
                   ROUND(AVG(cal.puntuacion), 1) as calificacion
            FROM solicitudes_servicio s
            INNER JOIN personas p ON s.id_trabajador = p.id_persona
            LEFT JOIN calificaciones cal ON s.id_solicitud = cal.id_solicitud
            WHERE s.estado = 'completada'
            GROUP BY s.id_trabajador, p.nombre_completo
            ORDER BY trabajos DESC LIMIT 10
        """)
        top_trabajadores = cursor.fetchall()

        # ── Solicitudes por ciudad ──
        cursor.execute("""
            SELECT ciudad, COUNT(*) as total FROM solicitudes_servicio
            WHERE ciudad IS NOT NULL AND ciudad != ''
            GROUP BY ciudad ORDER BY total DESC LIMIT 10
        """)
        por_ciudad = cursor.fetchall()

        cursor.close()
        conexion.close()

        # ── Crear Excel ──
        wb = openpyxl.Workbook()
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Hoja 1: Resumen
        ws = wb.active
        ws.title = "Resumen"
        resumen = [
            ["Métrica", "Valor"],
            ["Total Profesionales", total_profesionales],
            ["Total Clientes", total_clientes],
            ["Total Solicitudes", total_solicitudes],
            ["Completadas", completadas],
            ["Canceladas", canceladas],
            ["Pendientes", pendientes],
            ["Tasa Conversión", f"{round(completadas/(completadas+canceladas)*100, 1) if (completadas+canceladas) > 0 else 0}%"],
            ["Tarifa Promedio (COP)", f"${int(tarifa_prom):,}"],
            ["Fecha Reporte", dt.now().strftime('%d/%m/%Y %H:%M')],
        ]
        for row in resumen:
            ws.append(row)
        for col in range(1, 3):
            ws.cell(row=1, column=col).fill = header_fill
            ws.cell(row=1, column=col).font = header_font
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

        # Hoja 2: Top Categorías
        ws2 = wb.create_sheet("Top Categorías")
        ws2.append(["Categoría", "Solicitudes"])
        ws2.cell(row=1, column=1).fill = header_fill; ws2.cell(row=1, column=1).font = header_font
        ws2.cell(row=1, column=2).fill = header_fill; ws2.cell(row=1, column=2).font = header_font
        for cat in top_categorias:
            ws2.append([cat['categoria'], cat['total']])
        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 12

        # Gráfica de barras — Top Categorías
        from openpyxl.chart import BarChart, PieChart, Reference
        if len(top_categorias) > 0:
            chart_cat = BarChart()
            chart_cat.type = "col"
            chart_cat.title = "Solicitudes por Categoría"
            chart_cat.y_axis.title = "Solicitudes"
            data_ref = Reference(ws2, min_col=2, min_row=1, max_row=len(top_categorias)+1)
            cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=len(top_categorias)+1)
            chart_cat.add_data(data_ref, titles_from_data=True)
            chart_cat.set_categories(cats_ref)
            chart_cat.width = 18
            chart_cat.height = 10
            ws2.add_chart(chart_cat, "D2")

        # Hoja 3: Top Trabajadores
        ws3 = wb.create_sheet("Top Trabajadores")
        ws3.append(["Nombre", "Trabajos Completados", "Calificación"])
        for col in range(1, 4):
            ws3.cell(row=1, column=col).fill = header_fill; ws3.cell(row=1, column=col).font = header_font
        for t in top_trabajadores:
            ws3.append([t['nombre_completo'], t['trabajos'], float(t['calificacion'] or 0)])
        ws3.column_dimensions['A'].width = 30
        ws3.column_dimensions['B'].width = 20
        ws3.column_dimensions['C'].width = 14

        # Hoja 4: Por Ciudad
        ws4 = wb.create_sheet("Por Ciudad")
        ws4.append(["Ciudad", "Solicitudes"])
        ws4.cell(row=1, column=1).fill = header_fill; ws4.cell(row=1, column=1).font = header_font
        ws4.cell(row=1, column=2).fill = header_fill; ws4.cell(row=1, column=2).font = header_font
        for c in por_ciudad:
            ws4.append([c['ciudad'], c['total']])
        ws4.column_dimensions['A'].width = 20
        ws4.column_dimensions['B'].width = 12

        # Gráfica de pastel — Por Ciudad
        if len(por_ciudad) > 0:
            chart_city = PieChart()
            chart_city.title = "Solicitudes por Ciudad"
            data_ref = Reference(ws4, min_col=2, min_row=1, max_row=len(por_ciudad)+1)
            cats_ref = Reference(ws4, min_col=1, min_row=2, max_row=len(por_ciudad)+1)
            chart_city.add_data(data_ref, titles_from_data=True)
            chart_city.set_categories(cats_ref)
            chart_city.width = 14
            chart_city.height = 10
            ws4.add_chart(chart_city, "D2")

        # Hoja 5: Resumen Visual (gráfica de estados)
        ws5 = wb.create_sheet("Estados")
        ws5.append(["Estado", "Cantidad"])
        ws5.cell(row=1, column=1).fill = header_fill; ws5.cell(row=1, column=1).font = header_font
        ws5.cell(row=1, column=2).fill = header_fill; ws5.cell(row=1, column=2).font = header_font
        estados_data = [
            ["Completadas", completadas],
            ["Canceladas", canceladas],
            ["Pendientes", pendientes],
        ]
        for row in estados_data:
            ws5.append(row)
        ws5.column_dimensions['A'].width = 16
        ws5.column_dimensions['B'].width = 12

        chart_estado = PieChart()
        chart_estado.title = "Distribución por Estado"
        data_ref = Reference(ws5, min_col=2, min_row=1, max_row=4)
        cats_ref = Reference(ws5, min_col=1, min_row=2, max_row=4)
        chart_estado.add_data(data_ref, titles_from_data=True)
        chart_estado.set_categories(cats_ref)
        chart_estado.width = 14
        chart_estado.height = 10
        ws5.add_chart(chart_estado, "D2")

        # Guardar
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        nombre = f"metricas_talenthub_{dt.now().strftime('%Y%m%d')}.xlsx"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={nombre}"}
        )
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/clientes", response_class=HTMLResponse)
def ver_clientes(request: Request):
    """Página para ver todos los clientes registrados"""
    return templates.TemplateResponse("trabajadores/clientes.html", {"request": request})

@router.get("/clientes/listar")
def listar_clientes(request: Request):
    """API para obtener todos los clientes registrados (solo admin)"""
    if not verificar_admin(request):
        return JSONResponse({"error": "No autorizado", "clientes": []}, status_code=401)
    conexion = conectar_bd()
    if not conexion:
        return JSONResponse({"error": "Error de conexión", "clientes": []}, status_code=500)
    try:
        cursor = conexion.cursor(dictionary=True)
        cursor.execute("""
            SELECT c.id_cliente, c.nombre_completo, c.estado, c.fecha_registro,
                   e.correo, t.telefono,
                   (SELECT COUNT(*) FROM solicitudes_servicio s WHERE s.id_cliente = c.id_cliente) AS total_solicitudes,
                   (SELECT ROUND(AVG(cal.puntuacion),1)
                    FROM calificaciones cal
                    WHERE cal.id_cliente = c.id_cliente
                      AND cal.tipo_calificacion = 'trabajador_a_cliente') AS calificacion_como_cliente
            FROM clientes c
            LEFT JOIN correo_cliente e ON c.id_cliente = e.id_cliente AND e.principal = 1
            LEFT JOIN telefono_cliente t ON c.id_cliente = t.id_cliente AND t.principal = 1
            WHERE c.estado != 'eliminado' OR c.estado IS NULL
            ORDER BY c.id_cliente DESC
        """)
        clientes = cursor.fetchall()
        from datetime import timedelta
        for cl in clientes:
            for k, v in cl.items():
                if hasattr(v, 'isoformat'):
                    cl[k] = (v - timedelta(hours=5)).strftime('%Y-%m-%d %H:%M')
                elif v is None:
                    cl[k] = ''
                elif hasattr(v, '__float__'):
                    cl[k] = float(v)
        total = len(clientes)
        return JSONResponse({"clientes": clientes, "total": total})
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e), "clientes": []}, status_code=500)
    finally:
        if conexion and conexion.is_connected():
            conexion.close()

@router.post("/clientes/eliminar")
def eliminar_cliente(id_cliente: int = Form(...)):
    """Desactiva un cliente (soft delete — no borra datos)"""
    conexion = conectar_bd()
    if not conexion:
        return JSONResponse({"error": "Error de conexión"}, status_code=500)
    try:
        cursor = conexion.cursor(dictionary=True)
        # Verificar que existe
        cursor.execute("SELECT id_cliente, nombre_completo FROM clientes WHERE id_cliente = %s", (id_cliente,))
        cl = cursor.fetchone()
        if not cl:
            return JSONResponse({"error": "Cliente no encontrado"}, status_code=404)
        # Soft delete — desactivar
        cursor.execute("UPDATE clientes SET estado = 'inactivo' WHERE id_cliente = %s", (id_cliente,))
        conexion.commit()
        return JSONResponse({"mensaje": f"Cliente '{cl['nombre_completo']}' desactivado correctamente"})
    except Exception as e:
        conexion.rollback()
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        if conexion and conexion.is_connected():
            conexion.close()

@router.post("/clientes/reactivar")
def reactivar_cliente(id_cliente: int = Form(...)):
    """Reactiva un cliente desactivado"""
    conexion = conectar_bd()
    if not conexion:
        return JSONResponse({"error": "Error de conexión"}, status_code=500)
    try:
        cursor = conexion.cursor(dictionary=True)
        cursor.execute("SELECT id_cliente, nombre_completo FROM clientes WHERE id_cliente = %s", (id_cliente,))
        cl = cursor.fetchone()
        if not cl:
            return JSONResponse({"error": "Cliente no encontrado"}, status_code=404)
        cursor.execute("UPDATE clientes SET estado = 'activo' WHERE id_cliente = %s", (id_cliente,))
        conexion.commit()
        return JSONResponse({"mensaje": f"Cliente '{cl['nombre_completo']}' reactivado correctamente"})
    except Exception as e:
        conexion.rollback()
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        if conexion and conexion.is_connected():
            conexion.close()

@router.get("/archivo/{id_persona}/{tipo}")
def servir_archivo(id_persona: int, tipo: str):
    """Sirve archivos guardados en la base de datos"""
    from fastapi.responses import Response as FastResponse
    conexion = conectar_bd()
    try:
        cursor = conexion.cursor(dictionary=True)
        campos = {
            'foto': ('foto_identificacion_data', 'foto_identificacion_tipo'),
            'antecedentes': ('antecedentes_data', 'antecedentes_tipo'),
            'recomendaciones': ('recomendaciones_data', 'recomendaciones_tipo'),
        }
        if tipo not in campos:
            return JSONResponse({"error": "Tipo inválido"}, status_code=400)
        
        col_data, col_tipo = campos[tipo]
        cursor.execute(f"SELECT {col_data}, {col_tipo} FROM detalles_persona WHERE id_persona = %s LIMIT 1", (id_persona,))
        row = cursor.fetchone()
        
        if not row or not row[col_data]:
            return JSONResponse({"error": "Archivo no encontrado"}, status_code=404)
        
        content_type = row[col_tipo] or 'application/octet-stream'
        return FastResponse(content=bytes(row[col_data]), media_type=content_type)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        if conexion and conexion.is_connected():
            conexion.close()


# ============================================
# GEOLOCALIZACIÓN Y DISPONIBILIDAD
# ============================================

@router.get("/disponibilidad", response_class=HTMLResponse)
def mostrar_disponibilidad(request: Request):
    """Página para que trabajadores activen/desactiven su disponibilidad"""
    return templates.TemplateResponse("trabajadores/mi_disponibilidad.html", {"request": request})

@router.get("/disponibilidad/estado")
def obtener_estado_disponibilidad(id_trabajador: int):
    """Obtiene el estado actual de disponibilidad del trabajador"""
    conexion = conectar_bd()
    if not conexion:
        return JSONResponse({"error": "Error de conexión"}, status_code=500)
    
    try:
        cursor = conexion.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT disponible, latitud, longitud, ultima_actualizacion
            FROM disponibilidad
            WHERE id_persona = %s
            LIMIT 1
        """, (id_trabajador,))
        
        result = cursor.fetchone()
        
        if result:
            # Convertir Decimal a float
            if result.get('latitud'):
                result['latitud'] = float(result['latitud'])
            if result.get('longitud'):
                result['longitud'] = float(result['longitud'])
            if result.get('ultima_actualizacion'):
                result['ultima_actualizacion'] = result['ultima_actualizacion'].isoformat()
            
            return JSONResponse(result)
        else:
            return JSONResponse({"disponible": False})
            
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()

@router.post("/disponibilidad/actualizar")
async def actualizar_disponibilidad(
    id_trabajador: int = Form(...),
    disponible: bool = Form(...),
    latitud: float = Form(None),
    longitud: float = Form(None)
):
    """Actualiza la disponibilidad y ubicación del trabajador"""
    conexion = conectar_bd()
    if not conexion:
        return JSONResponse({"error": "Error de conexión"}, status_code=500)
    
    try:
        cursor = conexion.cursor()
        
        # Verificar si ya existe registro
        cursor.execute("""
            SELECT id_disponibilidad FROM disponibilidad
            WHERE id_persona = %s
            LIMIT 1
        """, (id_trabajador,))
        
        existe = cursor.fetchone()
        
        if existe:
            # Actualizar
            cursor.execute("""
                UPDATE disponibilidad
                SET disponible = %s, latitud = %s, longitud = %s,
                    ultima_actualizacion = NOW()
                WHERE id_persona = %s
            """, (disponible, latitud, longitud, id_trabajador))
        else:
            # Insertar nuevo
            cursor.execute("""
                INSERT INTO disponibilidad (id_persona, disponible, latitud, longitud)
                VALUES (%s, %s, %s, %s)
            """, (id_trabajador, disponible, latitud, longitud))
        
        conexion.commit()
        
        return JSONResponse({
            "mensaje": "Disponibilidad actualizada correctamente",
            "disponible": disponible,
            "latitud": latitud,
            "longitud": longitud
        })
        
    except Exception as e:
        conexion.rollback()
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()


@router.get("/eliminados", response_class=HTMLResponse)
def ver_eliminados(request: Request):
    """Página de papelera de registros eliminados (solo admin)"""
    redir = requiere_admin(request)
    if redir: return redir
    return templates.TemplateResponse("trabajadores/eliminados.html", {"request": request})

@router.get("/eliminados/listar")
def listar_eliminados():
    """API para obtener registros eliminados"""
    conexion = conectar_bd()
    if not conexion:
        return JSONResponse({"error": "Error de conexión"}, status_code=500)
    
    try:
        cursor = conexion.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT p.id_persona, p.numero_documento, p.nombre_completo, p.ciudad, 
                   p.fecha_eliminacion, p.departamento,
                   tp.telefono
            FROM personas p
            LEFT JOIN telefono_persona tp ON p.id_persona = tp.id_persona
            WHERE p.estado = 'eliminado'
            ORDER BY p.fecha_eliminacion DESC
        """)
        
        registros = cursor.fetchall()
        
        # Serializar fechas
        for r in registros:
            for k, v in r.items():
                if hasattr(v, 'isoformat'):
                    from datetime import timedelta
                    v_col = v - timedelta(hours=5)
                    r[k] = v_col.strftime('%Y-%m-%d %H:%M')
                elif v is None:
                    r[k] = ''
            r['tipo_documento'] = 'CC'
        
        cursor.execute("SELECT COUNT(*) as total FROM personas WHERE estado = 'eliminado'")
        total = cursor.fetchone()['total']
        
        return JSONResponse({
            "registros": registros,
            "total": total
        })
        
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        if conexion and conexion.is_connected():
            conexion.close()

@router.post("/restaurar")
async def restaurar_registro(id_persona: int = Form(...)):
    """Restaurar un registro eliminado"""
    conexion = conectar_bd()
    if not conexion:
        return JSONResponse({"error": "Error de conexión"}, status_code=500)
    
    try:
        cursor = conexion.cursor()
        
        cursor.execute("""
            UPDATE personas 
            SET estado = 'activo', fecha_eliminacion = NULL
            WHERE id_persona = %s
        """, (id_persona,))
        
        conexion.commit()
        
        return JSONResponse({"mensaje": "✅ Registro restaurado exitosamente"})
        
    except Exception as e:
        conexion.rollback()
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        if conexion and conexion.is_connected():
            conexion.close()

@router.post("/eliminar-permanente")
async def eliminar_permanente(id_persona: int = Form(...)):
    """Eliminar registro permanentemente de la base de datos"""
    import os

    conexion = conectar_bd()
    if not conexion:
        return JSONResponse({"error": "Error de conexión"}, status_code=500)

    try:
        cursor = conexion.cursor()

        # Obtener archivos antes de eliminar
        cursor.execute("""
            SELECT foto_identificacion, antecedentes_pdf, recomendaciones_archivo
            FROM detalles_persona
            WHERE id_persona = %s
        """, (id_persona,))
        archivos = cursor.fetchone()

        # Eliminar registros relacionados en orden para respetar foreign keys
        tablas_relacionadas = [
            "DELETE FROM push_subscriptions   WHERE tipo_usuario = 'trabajador' AND id_usuario = %s",
            "DELETE FROM mensajes_chat        WHERE id_remitente = %s AND tipo_remitente = 'trabajador'",
            "DELETE FROM calificaciones        WHERE id_trabajador = %s",
            "UPDATE solicitudes_servicio       SET id_trabajador = NULL WHERE id_trabajador = %s",
            "DELETE FROM disponibilidad        WHERE id_persona = %s",
            "DELETE FROM servicios_persona     WHERE id_persona = %s",
            "DELETE FROM experiencia_persona   WHERE id_persona = %s",
            "DELETE FROM telefono_persona      WHERE id_persona = %s",
            "DELETE FROM correo_persona        WHERE id_persona = %s",
            "DELETE FROM detalles_persona      WHERE id_persona = %s",
            "DELETE FROM sesiones              WHERE tipo_usuario = 'trabajador' AND id_usuario = %s",
            "DELETE FROM codigos_verificacion  WHERE tipo_usuario = 'trabajador' AND id_usuario = %s",
            "DELETE FROM tokens_recuperacion   WHERE tipo_usuario = 'trabajador' AND id_usuario = %s",
            "DELETE FROM trabajador_categorias  WHERE id_trabajador = %s",
        ]

        for sql in tablas_relacionadas:
            try:
                cursor.execute(sql, (id_persona,))
            except Exception:
                pass  # Si la tabla no existe, continuar

        # Ahora sí eliminar la persona
        cursor.execute("DELETE FROM personas WHERE id_persona = %s", (id_persona,))
        conexion.commit()

        # Eliminar archivos físicos
        if archivos:
            UPLOAD_FOLDER = "static/uploads"
            for archivo in archivos:
                if archivo:
                    archivo_path = os.path.join(UPLOAD_FOLDER, archivo)
                    if os.path.exists(archivo_path):
                        os.remove(archivo_path)

        return JSONResponse({"mensaje": "✅ Registro eliminado permanentemente"})

    except Exception as e:
        conexion.rollback()
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        if conexion and conexion.is_connected():
            conexion.close()


# ============================================
# AUTENTICACIÓN - LOGIN Y VERIFICACIÓN
# ============================================

@router.get("/login", response_class=HTMLResponse)
def mostrar_login_trabajador(request: Request):
    """Muestra página de login para trabajadores"""
    return templates.TemplateResponse("trabajadores/login.html", {"request": request})

@router.post("/login")
async def login_trabajador(
    response: Response,
    numero_documento: str = Form(...),
    password: str = Form(...)
):
    """Login de trabajador con número de documento + contraseña"""
    try:
        conexion = conectar_bd()
        cursor = conexion.cursor(dictionary=True)
        cursor.execute("""
            SELECT p.*, tp.telefono
            FROM personas p
            LEFT JOIN telefono_persona tp ON p.id_persona = tp.id_persona
            WHERE p.numero_documento = %s AND (p.estado = 'activo' OR p.estado IS NULL)
            LIMIT 1
        """, (numero_documento,))
        trabajador = cursor.fetchone()
        cursor.close()
        conexion.close()

        if not trabajador:
            return JSONResponse({"error": "Documento no registrado"}, status_code=401)
        if not trabajador.get('password_hash'):
            return JSONResponse({"error": "Debes crear tu contraseña primero", "redirect": f"/trabajador/crear_password?id_persona={trabajador['id_persona']}"}, status_code=401)
        if not auth.verificar_password(password, trabajador['password_hash']):
            return JSONResponse({"error": "Contraseña incorrecta"}, status_code=401)

        token = auth.crear_sesion('trabajador', trabajador['id_persona'])
        response.set_cookie(
            key="session_token",
            value=token,
            httponly=False,  # False para que JS pueda leerla si es necesario
            max_age=86400,
            samesite="lax"
        )
        nombre = trabajador['nombre_completo'].split()[0] if trabajador.get('nombre_completo') else 'Trabajador'
        return JSONResponse({
            "mensaje": f"Bienvenido, {nombre}",
            "redirect": f"/trabajador/panel?nombre={trabajador['nombre_completo']}&doc={trabajador['numero_documento']}&id={trabajador['id_persona']}"
        })

    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@router.post("/verificar-sms")
async def verificar_sms_trabajador(
    response: Response,
    id_trabajador: int = Form(...),
    codigo: str = Form(...)
):
    """Verifica el código SMS y crea la sesión"""
    try:
        valido, mensaje = auth.verificar_codigo_sms('trabajador', id_trabajador, codigo)
        
        if not valido:
            return JSONResponse(
                {"error": mensaje},
                status_code=400
            )
        
        # Crear sesión
        token = auth.crear_sesion('trabajador', id_trabajador)
        
        # Guardar token en cookie
        response.set_cookie(
            key="session_token",
            value=token,
            httponly=True,
            max_age=86400,
            samesite="lax"
        )
        
        return JSONResponse({
            "mensaje": "Verificación exitosa",
            "redirect": "/trabajador/disponibilidad"
        })
        
    except Exception as e:
        return JSONResponse(
            {"error": f"Error: {str(e)}"},
            status_code=500
        )

@router.post("/reenviar-codigo")
async def reenviar_codigo_trabajador(id_trabajador: int = Form(...)):
    """Reenvía el código SMS"""
    try:
        conexion = conectar_bd()
        cursor = conexion.cursor(dictionary=True)
        
        # Obtener teléfono del trabajador
        cursor.execute("""
            SELECT t.telefono
            FROM telefono_persona t
            WHERE t.id_persona = %s
            LIMIT 1
        """, (id_trabajador,))
        
        result = cursor.fetchone()
        cursor.close()
        conexion.close()
        
        if not result:
            return JSONResponse(
                {"error": "No se encontró teléfono registrado"},
                status_code=404
            )
        
        # Generar nuevo código
        codigo = auth.crear_codigo_verificacion(
            'trabajador',
            id_trabajador,
            result['telefono'],
            'login'
        )
        
        # Enviar SMS
        auth.enviar_sms(result['telefono'], codigo)
        
        return JSONResponse({"mensaje": "Código reenviado"})
        
    except Exception as e:
        return JSONResponse(
            {"error": f"Error: {str(e)}"},
            status_code=500
        )

@router.get("/logout")
def logout_trabajador(response: Response):
    """Cierra la sesión del trabajador"""
    response.delete_cookie("session_token")
    return RedirectResponse(url="/", status_code=302)

@router.post("/configurar-password")
async def configurar_password_trabajador(
    id_persona: int = Form(...),
    password: str = Form(...),
    confirmar_password: str = Form(...)
):
    """Configura la contraseña de un trabajador recién registrado"""
    try:
        if password != confirmar_password:
            return JSONResponse(
                {"error": "Las contraseñas no coinciden"},
                status_code=400
            )
        
        if len(password) < 6:
            return JSONResponse(
                {"error": "La contraseña debe tener al menos 6 caracteres"},
                status_code=400
            )
        
        auth.configurar_password_trabajador(id_persona, password)
        
        return JSONResponse({
            "mensaje": "Contraseña configurada correctamente",
            "redirect": "/trabajador/login"
        })
        
    except Exception as e:
        return JSONResponse(
            {"error": f"Error: {str(e)}"},
            status_code=500
        )

# ============================================
# RECUPERACIÓN DE CONTRASEÑA - TRABAJADORES
# ============================================

@router.get("/recuperar", response_class=HTMLResponse)
def mostrar_recuperar_trabajador(request: Request):
    """Página para solicitar recuperación de contraseña"""
    return templates.TemplateResponse("trabajadores/recuperar_password.html", {"request": request})

@router.post("/recuperar/solicitar")
async def solicitar_recuperacion_trabajador(request: Request, correo: str = Form(...)):
    """Genera token y envía email de recuperación"""
    try:
        token = auth.crear_token_recuperacion('trabajador', correo)
        if token:
            base_url = str(request.base_url).rstrip('/')
            # Enviar en background para no bloquear la respuesta
            import threading
            def _enviar():
                auth.enviar_email_recuperacion(correo, token, 'trabajador', base_url)
            threading.Thread(target=_enviar, daemon=True).start()
        # Siempre responder igual — no revelar si el correo existe o no
        return JSONResponse({"mensaje": "Si el correo está registrado, recibirás un enlace en breve."})
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": f"Error interno: {str(e)}"}, status_code=500)

@router.get("/recuperar/nueva-password", response_class=HTMLResponse)
def mostrar_nueva_password_trabajador(request: Request, token: str = ""):
    """Página para ingresar la nueva contraseña"""
    datos = auth.verificar_token_recuperacion(token) if token else None
    if not datos:
        return templates.TemplateResponse("trabajadores/recuperar_password.html", {
            "request": request,
            "error_token": "El enlace es inválido o ya expiró. Solicita uno nuevo."
        })
    return templates.TemplateResponse("trabajadores/nueva_password.html", {
        "request": request, "token": token
    })

@router.post("/recuperar/nueva-password")
async def guardar_nueva_password_trabajador(
    token: str = Form(...),
    password: str = Form(...),
    confirmar_password: str = Form(...)
):
    """Guarda la nueva contraseña del trabajador"""
    if password != confirmar_password:
        return JSONResponse({"error": "Las contraseñas no coinciden"}, status_code=400)
    if len(password) < 6:
        return JSONResponse({"error": "Mínimo 6 caracteres"}, status_code=400)
    ok = auth.consumir_token_recuperacion(token, password)
    if not ok:
        return JSONResponse({"error": "El enlace es inválido o ya expiró"}, status_code=400)
    return JSONResponse({"mensaje": "Contraseña actualizada correctamente", "redirect": "/trabajador/login"})

@router.get("/recuperar/verificar-correo")
def verificar_correo_trabajador(correo: str):
    """Verifica si un correo está registrado como trabajador (para validación en tiempo real)"""
    conexion = conectar_bd()
    try:
        cursor = conexion.cursor(dictionary=True)
        cursor.execute("""
            SELECT p.id_persona FROM personas p
            INNER JOIN correo_persona cp ON p.id_persona = cp.id_persona
            WHERE cp.correo = %s AND (p.estado = 'activo' OR p.estado IS NULL)
            LIMIT 1
        """, (correo,))
        existe = cursor.fetchone() is not None
        return JSONResponse({"existe": existe})
    except Exception as e:
        return JSONResponse({"existe": False, "error": str(e)})
    finally:
        if conexion and conexion.is_connected():
            conexion.close()

# ============================================
# VERIFICACIÓN DE EMAIL
# ============================================

@router.get("/verificar-email", response_class=HTMLResponse)
def verificar_email_trabajador(request: Request, token: str = ""):
    """Verifica el email del trabajador desde el link del correo"""
    if not token:
        return templates.TemplateResponse("verificacion_email.html", {
            "request": request, "exito": False,
            "mensaje": "Enlace inválido. Asegúrate de copiar el link completo del correo.",
            "redirect": "/trabajador/login"
        })
    datos = auth.verificar_email_token(token)
    if datos:
        return templates.TemplateResponse("verificacion_email.html", {
            "request": request, "exito": True,
            "mensaje": "¡Tu correo ha sido verificado exitosamente! Ya puedes acceder a todas las funciones de TalentHub.",
            "redirect": "/trabajador/login"
        })
    return templates.TemplateResponse("verificacion_email.html", {
        "request": request, "exito": False,
        "mensaje": "El enlace es inválido o ya expiró. Inicia sesión para solicitar uno nuevo.",
        "redirect": "/trabajador/login"
    })

@router.post("/reenviar-verificacion")
async def reenviar_verificacion_trabajador(request: Request):
    """Reenvía el email de verificación al trabajador autenticado"""
    import os
    token = request.cookies.get("session_token")
    if not token:
        return JSONResponse({"error": "No autenticado"}, status_code=401)
    sesion = auth.verificar_sesion(token)
    if not sesion or sesion['tipo_usuario'] != 'trabajador':
        return JSONResponse({"error": "Sesión inválida"}, status_code=401)

    conexion = conectar_bd()
    try:
        cursor = conexion.cursor(dictionary=True)
        cursor.execute("""
            SELECT cp.correo, cp.verificado, p.nombre_completo
            FROM correo_persona cp
            INNER JOIN personas p ON cp.id_persona = p.id_persona
            WHERE cp.id_persona = %s LIMIT 1
        """, (sesion['id_usuario'],))
        row = cursor.fetchone()
        if not row:
            return JSONResponse({"error": "No se encontró correo registrado"}, status_code=404)
        if row['verificado']:
            return JSONResponse({"mensaje": "Tu correo ya está verificado."})

        base_url = os.getenv("APP_URL", "https://web-production-191f4.up.railway.app")
        import threading
        resultado = [False]
        def _enviar():
            resultado[0] = auth.enviar_email_bienvenida(
                row['correo'], row['nombre_completo'], 'trabajador',
                sesion['id_usuario'], base_url
            )
        t = threading.Thread(target=_enviar, daemon=True)
        t.start()
        t.join(timeout=15)
        if resultado[0]:
            return JSONResponse({"mensaje": f"Correo de verificación reenviado a {row['correo']}"})
        else:
            return JSONResponse({"error": "No se pudo enviar el correo. Revisa GMAIL_USER y GMAIL_PASS en Railway."}, status_code=500)
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        if conexion and conexion.is_connected():
            conexion.close()
